import sys
from PyQt6.QtWidgets import (QMainWindow, QVBoxLayout, QWidget, QLabel, QHBoxLayout, 
                             QFrame, QGraphicsDropShadowEffect, QApplication, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QComboBox, QPushButton, QDialog,
                             QLineEdit, QFormLayout, QMessageBox, QFileDialog, QAbstractItemView, QScrollArea,
                             QStackedWidget, QSizePolicy, QDateEdit)
from PyQt6.QtCore import Qt, QSize, pyqtSignal, QTimer, QDate, QThread
from PyQt6.QtGui import QColor, QFont, QIcon
from database.db_handler import (get_consolidated_stats, get_recent_transactions, 
                                 get_top_items, add_expense, get_analytics_data, 
                                 CURRENT_SHOP_ID, set_shop_id, export_to_excel,
                                 get_history_data, get_dashboard_insights, get_app_settings, get_notifications,
                                 get_chart_data, get_recent_stock_alerts, get_report_data)
from datetime import datetime
import pandas as pd
import os
from ui.admin_settings_window import AdminSettingsView
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from ui.styles import STYLE_SHEET

class ExpenseDialog(QDialog):
    data_updated = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Shop Expense")
        self.setFixedSize(400, 350)
        self.setStyleSheet("""
            QDialog { background-color: #F8F9FA; font-family: 'Segoe UI'; }
            QLabel { color: #2D3436; font-weight: bold; }
            QLineEdit, QComboBox {
                background-color: white;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                padding: 8px;
                color: #2D3436;
            }
            QPushButton {
                background-color: #1b4d89;
                color: white;
                border-radius: 8px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover { background-color: #245fa8; }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30,30,30,30)
        
        title = QLabel("Add New Expense")
        title.setStyleSheet("font-size: 20px; margin-bottom: 20px;")
        layout.addWidget(title)
        
        form = QFormLayout()
        form.setSpacing(15)
        
        self.cat_input = QComboBox()
        self.cat_input.addItems(["Utility Bills", "Rent", "Staff Salary", "Maintenance", "Others"])
        
        self.amount_input = QLineEdit()
        self.amount_input.setPlaceholderText("Enter amount (Rs.)")
        
        self.desc_input = QLineEdit()
        self.desc_input.setPlaceholderText("Short description")
        
        form.addRow("Category:", self.cat_input)
        form.addRow("Amount:", self.amount_input)
        form.addRow("Description:", self.desc_input)
        
        layout.addLayout(form)
        layout.addSpacing(20)
        
        self.save_btn = QPushButton("Save Expense")
        self.save_btn.clicked.connect(self.handle_save)
        layout.addWidget(self.save_btn)

    def handle_save(self):
        cat = self.cat_input.currentText()
        amt = self.amount_input.text()
        desc = self.desc_input.text()
        if not amt or not desc:
            QMessageBox.warning(self, "Missing Info", "Please fill all fields.")
            return
        try:
            amount_val = float(amt)
            add_expense(cat, amount_val, desc)
            self.data_updated.emit()
            self.accept()
        except ValueError:
            QMessageBox.warning(self, "Invalid Amount", "Please enter a valid numeric value for the amount.")

class ReportWorker(QThread):
    report_done = pyqtSignal(str)
    report_failed = pyqtSignal(str)

    def __init__(self, from_date, to_date, format_type):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
        self.format_type = format_type

    def run(self):
        try:
            from database.db_handler import get_report_data
            from utils.report_generator import generate_pdf_report, generate_excel_report
            
            data = get_report_data(self.from_date, self.to_date)
            
            if self.format_type == "PDF":
                filepath = generate_pdf_report(data, self.from_date, self.to_date)
            else:
                filepath = generate_excel_report(data, self.from_date, self.to_date)
            
            if filepath:
                self.report_done.emit(filepath)
            else:
                self.report_failed.emit("Failed to generate report file.")
        except Exception as e:
            self.report_failed.emit(str(e))

class AdminDashboard(QMainWindow):
    logout_clicked = pyqtSignal()
    settings_updated = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Admin Control Panel")
        self.showMaximized()
        
        # --- GLOBAL UI STYLE ---
        self.setStyleSheet(STYLE_SHEET)
        
        self.central_widget = QWidget()
        self.central_widget.setObjectName("AdminMain")
        self.central_widget.setStyleSheet("QWidget#AdminMain { background-color: #F0F2F5; }")
        self.setCentralWidget(self.central_widget)
        
        # Horizontal layout to split Sidebar and Content
        self.root_layout = QHBoxLayout(self.central_widget)
        self.root_layout.setContentsMargins(0, 0, 0, 0)
        self.root_layout.setSpacing(0)
        
        # --- 1. SIDEBAR ---
        self.sidebar = QFrame()
        self.sidebar.setFixedWidth(270)
        self.sidebar.setObjectName("AdminSidebar")
        self.sidebar.setStyleSheet("""
            QFrame#AdminSidebar { 
                background-color: #0F172A; /* Deep Navy/Slate 900 */
            }
            QPushButton {
                text-align: left;
                padding: 14px 18px;
                border: none;
                border-radius: 12px;
                background-color: transparent;
                color: #94A3B8; /* Muted Slate */
                font-weight: 600;
                font-size: 14px;
                margin: 4px 15px;
            }
            QPushButton:hover { 
                background-color: rgba(255, 255, 255, 0.05); 
                color: #FFFFFF; 
            }
            QPushButton#ActiveNav { 
                background-color: #3B82F6; /* Primary Modern Blue */
                color: #FFFFFF; 
                font-weight: 700;
                border-radius: 12px;
            }
        """)
        
        side_layout = QVBoxLayout(self.sidebar)
        side_layout.setContentsMargins(0, 25, 0, 25)
        side_layout.setSpacing(8)
        
        # Premium Brand Badge
        brand_area = QFrame()
        brand_area.setStyleSheet("margin: 0px 20px 20px 20px; padding: 15px; background: rgba(255,255,255,0.03); border-radius: 20px; border: 1px solid rgba(255,255,255,0.05);")
        brand_layout = QHBoxLayout(brand_area)
        brand_layout.setContentsMargins(10, 5, 10, 5)
        
        brand_icon = QLabel("⚡")
        brand_icon.setStyleSheet("font-size: 22px; background: transparent;")
        
        shop_name = get_app_settings().get("shop_name", "Kamran Electronics")
        brand_text_layout = QVBoxLayout()
        brand_name = QLabel(shop_name)
        brand_name.setStyleSheet("color: #F8FAFC; font-size: 15px; font-weight: 800; background: transparent;")
        brand_tag = QLabel("ADMIN PANEL")
        brand_tag.setStyleSheet("color: #3B82F6; font-size: 9px; font-weight: 700; letter-spacing: 1px; background: transparent;")
        
        brand_text_layout.addWidget(brand_name)
        brand_text_layout.addWidget(brand_tag)
        
        brand_layout.addWidget(brand_icon)
        brand_layout.addLayout(brand_text_layout)
        brand_layout.addStretch()
        
        side_layout.addWidget(brand_area)

        # Nav Title
        nav_section_lbl = QLabel("MAIN MENU")
        nav_section_lbl.setStyleSheet("color: #475569; font-size: 11px; font-weight: 800; letter-spacing: 2px; padding: 10px 30px; margin-top: 10px;")
        side_layout.addWidget(nav_section_lbl)
        
        self.nav_home = QPushButton("🏠   Dashboard")
        self.nav_home.setObjectName("ActiveNav")
        self.nav_home.clicked.connect(lambda: self.switch_view(0, self.nav_home))
        
        self.nav_history = QPushButton("📜   History")
        self.nav_history.clicked.connect(lambda: self.switch_view(1, self.nav_history))
        
        self.nav_reports = QPushButton("📋   Reports")
        self.nav_reports.clicked.connect(lambda: self.switch_view(4, self.nav_reports))
        
        self.nav_staff = QPushButton("👥   Staff Management")
        self.nav_staff.clicked.connect(lambda: self.switch_view(2, self.nav_staff))
        
        self.nav_settings = QPushButton("⚙️   System Settings")
        self.nav_settings.clicked.connect(lambda: self.switch_view(3, self.nav_settings))
        
        side_layout.addWidget(self.nav_home)
        side_layout.addWidget(self.nav_history)
        side_layout.addWidget(self.nav_reports)
        side_layout.addWidget(self.nav_staff)
        side_layout.addWidget(self.nav_settings)
        side_layout.addStretch()
        
        # Sidebar Footer
        footer_card = QFrame()
        footer_card.setStyleSheet("background: rgba(15, 23, 42, 0.5); border-top: 1px solid rgba(255,255,255,0.05); padding: 20px;")
        footer_layout = QVBoxLayout(footer_card)
        
        version_lbl = QLabel("Build v1.0.4 - STABLE")
        version_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        version_lbl.setStyleSheet("color: #475569; font-size: 10px; font-weight: 600;")
        
        copyright_lbl = QLabel("© 2026 Kamran Electronics")
        copyright_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        copyright_lbl.setStyleSheet("color: #1E293B; font-size: 9px; margin-top: 2px;")
        
        footer_layout.addWidget(version_lbl)
        footer_layout.addWidget(copyright_lbl)
        
        side_layout.addWidget(footer_card)
        
        self.root_layout.addWidget(self.sidebar)
        
        # --- 2. MAIN CONTENT AREA ---
        self.content_area = QWidget()
        self.main_layout = QVBoxLayout(self.content_area)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        
        # --- GLOBAL SHARED NAVBAR (Persistent) ---
        self.setup_shared_navbar()
        
        # --- MARGINED CONTAINER FOR STACK ---
        self.stack_wrapper = QWidget()
        self.stack_wrapper_layout = QVBoxLayout(self.stack_wrapper)
        self.stack_wrapper_layout.setContentsMargins(0, 0, 0, 0) # Wrapper doesn't need margins if dash_layout handles it
        self.stack_wrapper_layout.setSpacing(0)
        
        # --- STACKED WIDGET FOR VIEWS ---
        self.stack = QStackedWidget()
        self.dashboard_container = QWidget()
        self.dash_layout = QVBoxLayout(self.dashboard_container)
        self.dash_layout.setContentsMargins(25, 15, 25, 25)
        self.dash_layout.setSpacing(12)
        
        self.stat_widgets = {} 
        
        # Remove top bar from here, it's global now
        self.setup_stats_cards()
        self.setup_center_section() # Implement Row 3
        self.setup_bottom_bar()
        
        self.stack.addWidget(self.dashboard_container) # Index 0
        
        # History View (Index 1)
        self.history_view = QWidget()
        self.history_view.setStyleSheet("background-color: white;")
        self.stack.addWidget(self.history_view)
        
        # Staff View (Index 2)
        from ui.staff_view import StaffView
        self.staff_container = StaffView()
        self.stack.addWidget(self.staff_container)
        
        # Settings View (Index 3)
        from ui.admin_settings_window import AdminSettingsView
        self.settings_view = AdminSettingsView()
        self.settings_view.back_clicked.connect(lambda: self.switch_view(0, self.nav_home))
        self.settings_view.settings_saved.connect(self.settings_updated.emit)
        self.stack.addWidget(self.settings_view)
        
        # Reports Page (Index 4)
        self.reports_page = QWidget()
        self.stack.addWidget(self.reports_page)
        
        self.stack_wrapper_layout.addWidget(self.stack)
        self.main_layout.addWidget(self.stack_wrapper)
        
        self.root_layout.addWidget(self.content_area)
        
        # Track buttons for ActiveNav style
        self.nav_buttons = [self.nav_home, self.nav_history, self.nav_reports, self.nav_staff, self.nav_settings]
        
        # Refresh now only AFTER UI is fully built
        self.refresh_dashboard('Today')
        self.update_notification_badge()
        
        # Auto refresh every 5 minutes
        self.auto_refresh_timer = QTimer()
        self.auto_refresh_timer.timeout.connect(self.auto_refresh)
        self.auto_refresh_timer.start(300000)

        # Build History Page Content
        self.build_history_page()
        
        # Build Reports Page Content
        self.build_reports_page()

    def switch_view(self, index, active_btn):
        """Switches the QStackedWidget index and handles ActiveNav button highlighting."""
        # Remove ActiveNav object name from all buttons
        for btn in self.nav_buttons:
            btn.setObjectName("")
            
        # Set ActiveNav to the clicked button
        active_btn.setObjectName("ActiveNav")
        
        # Force stylesheet update
        for btn in self.nav_buttons:
            btn.style().unpolish(btn)
            btn.style().polish(btn)
            
        self.stack.setCurrentIndex(index)
        
        if hasattr(self, 'nav_title'):
             if index == 1:
                 self.nav_title.setText("Transaction History")
             elif index == 4:
                 self.nav_title.setText("Reports")
             elif index == 2:
                 self.nav_title.setText("Staff Management")
             else:
                 self.nav_title.setText(active_btn.text().split(' ', 1)[-1])

        if hasattr(self, 'navbar_subtitle'):
            if index == 1:
                self.navbar_subtitle.setText("Complete record of all transactions")
            elif index == 4:
                self.navbar_subtitle.setText("Generate and export business reports")
            elif index == 2:
                self.navbar_subtitle.setText("Manage your team members")
             
        if index == 0:
            current_period = self.period_dropdown.currentText()
            self.refresh_dashboard(current_period)
        elif index == 2:
            self.staff_container.load_staff_data()


    def refresh_dashboard(self, time_period):
        try:
            # Map dropdown text to db function parameter
            period_map = {
                'Today': 'Today',
                'Weekly': 'Last Week',
                'Monthly': 'Last Month',
                'Overall': 'All Time'
            }
            db_period = period_map.get(time_period, 'Today')
            
            # Fetch fresh stats
            stats = get_consolidated_stats(CURRENT_SHOP_ID, db_period)
            
            # Update Row 1 cards
            revenue = stats.get('total_sales', 0.0)
            expenses = stats.get('total_expenses', 0.0)
            self.revenue_value_lbl.setText(f"Rs. {revenue:,.0f}")
            self.expense_value_lbl.setText(f"Rs. {expenses:,.0f}")
            
            # Update Row 2 cards
            product_sales = stats.get('total_sales', 0.0) - stats.get('repair_revenue', 0.0)
            repair_income = stats.get('repair_revenue', 0.0)
            net_profit = stats.get('net_profit', 0.0)
            
            self.product_sales_lbl.setText(f"Rs. {max(product_sales, 0):,.0f}")
            self.repair_income_lbl.setText(f"Rs. {repair_income:,.0f}")
            
            if net_profit >= 0:
                self.net_profit_lbl.setText(f"▲  Rs. {net_profit:,.0f}")
                self.net_profit_lbl.setStyleSheet("font-size: 20px; font-weight: 900; color: #10B981;")
            else:
                self.net_profit_lbl.setText(f"▼  Rs. {abs(net_profit):,.0f}")
                self.net_profit_lbl.setStyleSheet("font-size: 20px; font-weight: 900; color: #EF4444;")
            
            # Update chart
            self.update_chart(time_period)
            
            # Update alerts panel
            self.build_alerts_panel()
            
            # Update navbar subtitle
            subtitle_map = {
                'Today': "Showing today's performance 📅",
                'Weekly': "Showing last 7 days 📅",
                'Monthly': "Showing last 30 days 📅",
                'Overall': "Showing all time data 📅"
            }
            if hasattr(self, 'navbar_subtitle'):
                self.navbar_subtitle.setText(subtitle_map.get(time_period, "Welcome back, Admin 👋"))
            
            # Also update tables
            self.refresh_all_dashboard_data(time_period)
                
        except Exception as e:
            print(f"Dashboard refresh error: {e}")

    def build_alerts_panel(self):
        alerts = get_recent_stock_alerts()
        
        # Clear existing items in scroll widget layout
        while self.alerts_layout.count():
            child = self.alerts_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        
        if not alerts:
            empty_lbl = QLabel("✅ All systems normal")
            empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_lbl.setStyleSheet("color: #10B981; font-size: 14px; font-weight: 600; padding: 20px;")
            self.alerts_layout.addWidget(empty_lbl)
            return
        
        for alert_type, name, detail in alerts:
            item = QFrame()
            if alert_type == 'low_stock':
                item.setStyleSheet("""
                    QFrame {
                        background: #FFF7ED;
                        border-left: 3px solid #F59E0B;
                        border-radius: 10px;
                        padding: 2px;
                    }
                """)
                icon = "⚠️"
            else:
                item.setStyleSheet("""
                    QFrame {
                        background: #F0FDF4;
                        border-left: 3px solid #10B981;
                        border-radius: 10px;
                        padding: 2px;
                    }
                """)
                icon = "✅"
            
            item_layout = QVBoxLayout(item)
            item_layout.setContentsMargins(12, 10, 12, 10)
            item_layout.setSpacing(2)
            
            name_lbl = QLabel(f"{icon} {name}")
            name_lbl.setStyleSheet("font-size: 13px; font-weight: 700; color: #1E293B; background: transparent;")
            
            detail_lbl = QLabel(detail)
            detail_lbl.setStyleSheet("font-size: 11px; color: #64748B; background: transparent;")
            
            item_layout.addWidget(name_lbl)
            item_layout.addWidget(detail_lbl)
            
            self.alerts_layout.addWidget(item)
        
        self.alerts_layout.addStretch()

    def update_chart(self, time_period):
        try:
            from matplotlib.ticker import FuncFormatter
            import numpy as np
            labels, sales_values, expense_values = get_chart_data(time_period)
            self.chart_axes.clear()
            
            # Title
            title_map = {
                'Today': "📈 Today's Sales Analytics",
                'Weekly': "📈 Weekly Performance Trend",
                'Monthly': "📈 Monthly Revenue Overview",
                'Overall': "📈 Life-Time Sales Analytics"
            }
            if hasattr(self, 'sales_chart_title'):
                self.sales_chart_title.setText(title_map.get(time_period, "📈 Sales Analytics"))

            # Background
            self.chart_axes.set_facecolor('#FAFBFF')
            self.chart_figure.patch.set_facecolor('#FAFBFF')
            
            # 3. Shaded Profit/Loss Zone (Green for profit, red for loss)
            x_range = range(len(labels))
            s_arr = np.array(sales_values)
            e_arr = np.array(expense_values)
            self.chart_axes.fill_between(x_range, sales_values, expense_values, where=(s_arr > e_arr), 
                                         interpolate=True, color='#10B981', alpha=0.06, label='Profit Zone')
            self.chart_axes.fill_between(x_range, sales_values, expense_values, where=(s_arr <= e_arr), 
                                         interpolate=True, color='#EF4444', alpha=0.06, label='Loss Zone')
            
            # Core Lines
            self.chart_axes.plot(labels, sales_values, color='#3A8DFF', linewidth=2.0, label="Revenue")
            self.chart_axes.plot(labels, expense_values, color='#EF4444', linewidth=2.0, label="Expenses")
            
            # 4. Average Revenue Line
            if sales_values:
                avg_rev = sum(sales_values) / len(sales_values)
                self.chart_axes.axhline(avg_rev, color='#3A8DFF', alpha=0.3, linewidth=1, linestyle='--')
                self.chart_axes.text(len(labels)-1, avg_rev, 'Avg Revenue', color='#3A8DFF', fontsize=7, 
                                     va='bottom', ha='right', alpha=0.5, fontweight='bold')

            # 1. Peak Annotations
            if sales_values and max(sales_values) > 0:
                mx_r = max(sales_values)
                mx_r_idx = sales_values.index(mx_r)
                self.chart_axes.annotate(f"Peak: Rs.{mx_r:,.0f}", xy=(mx_r_idx, mx_r), xytext=(0, 10),
                                         textcoords='offset points', ha='center', fontsize=8, color='#3A8DFF',
                                         fontweight='bold', arrowprops=dict(arrowstyle='->', color='#3A8DFF', lw=0.5))
                self.chart_axes.plot(mx_r_idx, mx_r, 'o', markersize=4, color='#3A8DFF')

            if expense_values and max(expense_values) > 0:
                mx_e = max(expense_values)
                mx_e_idx = expense_values.index(mx_e)
                self.chart_axes.annotate(f"High Expense: Rs.{mx_e:,.0f}", xy=(mx_e_idx, mx_e), xytext=(0, 10),
                                         textcoords='offset points', ha='center', fontsize=8, color='#EF4444',
                                         fontweight='bold', arrowprops=dict(arrowstyle='->', color='#EF4444', lw=0.5))
                self.chart_axes.plot(mx_e_idx, mx_e, 'o', markersize=4, color='#EF4444')

            # 2. Summary Stats Bar (Top Left axes text)
            total_rev = sum(sales_values)
            total_exp = sum(expense_values)
            net_val = total_rev - total_exp
            stats_str = f"Total Revenue: Rs.{total_rev:,.0f}  |  Total Expenses: Rs.{total_exp:,.0f}  |  Net: Rs.{net_val:,.0f}"
            self.chart_axes.text(0.02, 0.96, stats_str, transform=self.chart_axes.transAxes, fontsize=8, 
                                 color='#64748B', fontweight='600', bbox=dict(facecolor='white', alpha=0.9, edgecolor='none', boxstyle='round,pad=0.4'))

            # Stylings
            indices = range(len(labels))
            visible_indices = [i for i in indices if i % 3 == 0]
            visible_labels = [labels[i] for i in visible_indices]
            self.chart_axes.set_xticks(visible_indices)
            self.chart_axes.set_xticklabels(visible_labels, rotation=0, fontsize=9, color='#94A3B8')
            
            self.chart_axes.yaxis.set_major_formatter(FuncFormatter(lambda x, p: f"Rs.{int(x):,}"))
            self.chart_axes.tick_params(axis='y', labelsize=9, labelcolor='#94A3B8')
            self.chart_axes.yaxis.grid(True, color='#EEF0F6', linewidth=0.6, linestyle='dashed')
            for spine in self.chart_axes.spines.values():
                spine.set_visible(False)
            
            self.chart_figure.tight_layout(pad=1.2)
            self.chart_canvas.draw()
        except Exception as e:
            print(f"Chart update error: {e}")

    def refresh_all_dashboard_data(self, period):
        """Master function to refresh stats, top items, and transaction history."""
        # Note: Stats are already refreshed in refresh_dashboard()
        # This can be used for any additional data needed specifically in dashboard view
        pass

    def event(self, event):
        from PyQt6.QtCore import QEvent
        if event.type() == QEvent.Type.WindowActivate:
            # Refresh using currently selected filter for FULL DATA (Charts + Tables)
            if hasattr(self, 'period_dropdown'):
                self.refresh_dashboard(self.period_dropdown.currentText())
        return super().event(event)

    def setup_shared_navbar(self):
        nav_frame = QFrame()
        nav_frame.setFixedHeight(70)
        nav_frame.setStyleSheet("""
            QFrame {
                background-color: #FFFFFF;
                border-bottom: 1px solid #E2E8F0;
                border-radius: 0px;
            }
        """)
        
        # Add shadow
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(8)
        shadow.setYOffset(2)
        shadow.setXOffset(0)
        shadow.setColor(QColor(0, 0, 0, 15)) # rgba(0,0,0,0.06) is ~15 in alpha
        nav_frame.setGraphicsEffect(shadow)
        
        nav_layout = QHBoxLayout(nav_frame)
        nav_layout.setContentsMargins(30, 0, 30, 0)
        
        # LEFT SIDE
        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(0, 10, 0, 10)
        left_layout.setSpacing(2)
        left_layout.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        
        self.nav_title = QLabel("Dashboard")
        self.nav_title.setStyleSheet("font-size: 20px; font-weight: 800; color: #1E293B; border: none; background: transparent;")
        
        self.navbar_subtitle = QLabel("Welcome back, Admin 👋")
        self.navbar_subtitle.setStyleSheet("font-size: 12px; color: #64748B; border: none; background: transparent;")
        
        left_layout.addWidget(self.nav_title)
        left_layout.addWidget(self.navbar_subtitle)
        
        nav_layout.addLayout(left_layout)
        nav_layout.addStretch()
        
        # RIGHT SIDE
        
        # Item 1: Dropdown
        self.period_dropdown = QComboBox()
        self.period_dropdown.addItems(["Today", "Weekly", "Monthly", "Overall"])
        self.period_dropdown.setMinimumWidth(150)
        self.period_dropdown.setCursor(Qt.CursorShape.PointingHandCursor)
        self.period_dropdown.setStyleSheet("""
            QComboBox {
                background-color: #FFFFFF;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                padding: 10px 40px 10px 20px;
                font-weight: 700;
                color: #1E293B;
                font-size: 13px;
            }
            QComboBox:hover {
                border-color: #3A8DFF;
                background-color: #F8FAFC;
            }
            QComboBox::drop-down {
                border: none;
                width: 40px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #64748B;
                margin-right: 20px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                selection-background-color: #F1F5F9;
                selection-color: #3A8DFF;
                padding: 8px;
            }
        """)
        # Subtle Shadow for Dropdown
        drop_shadow = QGraphicsDropShadowEffect()
        drop_shadow.setBlurRadius(10)
        drop_shadow.setYOffset(2)
        drop_shadow.setColor(QColor(0, 0, 0, 15))
        self.period_dropdown.setGraphicsEffect(drop_shadow)
        self.period_dropdown.currentTextChanged.connect(self.on_period_changed)
        nav_layout.addWidget(self.period_dropdown)
        nav_layout.addSpacing(15)
        
        # Item 2: Notification Bell
        self.btn_notif = QPushButton("🔔")
        self.btn_notif.setFixedSize(42, 42)
        self.btn_notif.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_notif.setStyleSheet("""
            QPushButton {
                background-color: #F8FAFC;
                border-radius: 21px;
                font-size: 18px;
                border: 1px solid #E2E8F0;
                color: #64748B;
            }
            QPushButton:hover { 
                background-color: #FFFFFF; 
                border-color: #3A8DFF;
                color: #3A8DFF;
            }
        """)
        self.btn_notif.clicked.connect(self.show_notifications)
        
        # Notification Badge
        self.notif_badge = QLabel(self.btn_notif)
        self.notif_badge.setFixedSize(10, 10)
        self.notif_badge.setStyleSheet("background-color: #EF4444; border-radius: 5px; border: 2px solid #FFFFFF;")
        self.notif_badge.move(28, 5) 
        self.notif_badge.hide()
        
        nav_layout.addWidget(self.btn_notif)
        nav_layout.addSpacing(15)
        
        # Item 3: Logout
        self.btn_logout = QPushButton("🚪 Logout")
        self.btn_logout.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_logout.setFixedHeight(42)
        self.btn_logout.setStyleSheet("""
            QPushButton {
                background-color: rgba(239, 68, 68, 0.08);
                color: #B91C1C;
                border-radius: 21px;
                padding: 0px 24px;
                font-weight: 700;
                border: 1px solid rgba(239, 68, 68, 0.15);
                font-size: 13px;
            }
            QPushButton:hover { 
                background-color: #EF4444; 
                color: #FFFFFF;
                border: none;
            }
        """)
        self.btn_logout.clicked.connect(self.handle_logout_click)
        nav_layout.addWidget(self.btn_logout)
        
        self.main_layout.addWidget(nav_frame)
        self.update_notification_badge()


    def build_history_page(self):
        """Builds the transaction history page content."""
        widget = self.stack.widget(1)
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(16)
        
        # Summary row
        self.history_summary_lbl = QLabel("Total Records: 0  |  Total Revenue: Rs. 0  |  Total Expenses: Rs. 0")
        self.history_summary_lbl.setStyleSheet("""
            background-color: #F8FAFF;
            border-radius: 10px;
            padding: 10px 20px;
            font-size: 12px;
            color: #64748B;
            border: 1px solid #E2E8F0;
        """)
        layout.addWidget(self.history_summary_lbl)

        # Search and filter row
        controls_layout = QHBoxLayout()
        controls_layout.setSpacing(12)
        
        self.history_search = QLineEdit()
        self.history_search.setPlaceholderText("Search by customer name or description...")
        self.history_search.setFixedHeight(45)
        self.history_search.setStyleSheet("""
            QLineEdit {
                background-color: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                padding: 10px 20px;
                font-size: 14px;
                color: #1E293B;
            }
            QLineEdit:focus {
                border-color: #3B82F6;
                background-color: #FFFFFF;
            }
        """)
        
        self.history_type_filter = QComboBox()
        self.history_type_filter.addItems(["All Types", "Sale", "Repair", "Expense", "Stock"])
        self.history_type_filter.setFixedHeight(45)
        self.history_type_filter.setMinimumWidth(150)
        self.history_type_filter.setStyleSheet("""
            QComboBox {
                background-color: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-radius: 12px;
                padding: 10px 20px;
                font-size: 14px;
                color: #1E293B;
            }
        """)
        
        controls_layout.addWidget(self.history_search, 4)
        controls_layout.addWidget(self.history_type_filter, 1)
        layout.addLayout(controls_layout)
        
        # Debounce Timer for search
        self.search_debounce_timer = QTimer()
        self.search_debounce_timer.setSingleShot(True)
        self.search_debounce_timer.timeout.connect(self.load_history_data)
        self.history_search.textChanged.connect(lambda: self.search_debounce_timer.start(300))
        
        self.history_type_filter.currentIndexChanged.connect(self.load_history_data)
        
        # History Table
        self.history_table = QTableWidget(0, 5)
        self.history_table.setHorizontalHeaderLabels(["Customer", "Type", "Description", "Amount", "Date"])
        self.history_table.verticalHeader().setVisible(False)
        self.history_table.horizontalHeader().setStretchLastSection(True)
        self.history_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.history_table.setShowGrid(False)
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setAlternatingRowColors(True)
        self.history_table.verticalHeader().setDefaultSectionSize(44)
        self.history_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 16px;
                border: 1px solid #E2E8F0;
                font-size: 13px;
                gridline-color: transparent;
                alternate-background-color: #FAFBFF;
            }
            QTableWidget::item {
                padding: 12px;
                border: none;
            }
            QTableWidget::item:hover {
                background-color: #F0F7FF;
            }
            QHeaderView::section {
                background-color: #FFFFFF;
                font-weight: 800;
                color: #1E293B;
                padding: 12px;
                border: none;
                border-bottom: 2px solid #F0F2F5;
            }
        """)
        
        layout.addWidget(self.history_table)
        
        # Initial data load
        self.load_history_data()

    def load_history_data(self):
        """Fetches and populates history data with local filtering."""
        period_text = self.period_dropdown.currentText()
        period_map = {
            'Today': 'Today',
            'Weekly': 'Last Week',
            'Monthly': 'Last Month',
            'Overall': 'All Time'
        }
        mapped_period = period_map.get(period_text, 'Today')
        
        # Fetch from DB
        results = get_history_data(mapped_period)
        
        search_text = self.history_search.text().strip().lower()
        type_filter = self.history_type_filter.currentText()
        
        # Filter results
        filtered_results = []
        for row in results:
            r_cust, r_type, r_desc, r_amt, r_date = row
            
            # Type filter
            if type_filter != "All Types" and r_type != type_filter:
                continue
                
            # Search filter
            if search_text and search_text not in r_desc.lower():
                continue
                
            filtered_results.append(row)
            
        # Clear and repopulate
        self.history_table.setRowCount(0)
        
        if not filtered_results:
            self.history_table.setRowCount(1)
            item = QTableWidgetItem("No records found for this period")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setForeground(QColor("#64748B"))
            self.history_table.setItem(0, 0, item)
            self.history_table.setSpan(0, 0, 1, 5)
            self.history_summary_lbl.setText("Total Records: 0  |  Total Revenue: Rs. 0  |  Total Expenses: Rs. 0")
            return

        total_rev = 0
        total_exp = 0
        
        for row_idx, row_data in enumerate(filtered_results):
            self.history_table.insertRow(row_idx)
            r_cust, r_type, r_desc, r_amt, r_date = row_data
            
            # Update summary counts
            if r_type in ["Sale", "Repair"]:
                total_rev += float(r_amt or 0)
            elif r_type == "Expense":
                total_exp += float(r_amt or 0)

            # Column 0: Customer
            cust_item = QTableWidgetItem(r_cust.title() if r_cust else "")
            cust_item.setForeground(QColor("#1E293B"))
            cust_item.setFont(QFont("Segoe UI", 10))
            cust_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row_idx, 0, cust_item)
            
            # Column 1: Type (Simplified Pills)
            badge_container = QWidget()
            badge_layout = QHBoxLayout(badge_container)
            badge_layout.setContentsMargins(0, 0, 0, 0)
            badge_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            badge = QLabel()
            badge_config = {
                "Sale": ("SALES", "#E0F2FE", "#0369A1"),
                "Repair": ("REPAIR", "#FEF3C7", "#92400E"),
                "Expense": ("EXPENSE", "#FEE2E2", "#991B1B"),
                "Stock": ("STOCK", "#DCFCE7", "#15803D")
            }
            type_label, bg, fg = badge_config.get(r_type, (r_type.upper(), "#F1F5F9", "#64748B"))
            display_text = f"● {type_label}"
            
            badge.setText(display_text)
            badge.setStyleSheet(f"""
                QLabel {{
                    background-color: {bg};
                    color: {fg};
                    border-radius: 10px;
                    padding: 4px 10px;
                    font-weight: 800;
                    font-size: 10px;
                    font-family: 'Segoe UI', sans-serif;
                }}
            """)
            badge_layout.addWidget(badge)
            self.history_table.setCellWidget(row_idx, 1, badge_container)
            
            # Column 2: Description
            desc_item = QTableWidgetItem(r_desc.title() if r_desc else "")
            desc_item.setForeground(QColor("#64748B"))
            desc_item.setFont(QFont("Segoe UI", 10))
            desc_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row_idx, 2, desc_item)
            
            # Column 3: Amount
            if r_type == "Stock":
                amt_str = f"{r_amt} Units"
                amt_color = "#3B82F6"
            else:
                amt_str = f"Rs. {r_amt:,.0f}"
                amt_color = "#EF4444" if r_type == "Expense" else "#10B981"
                
            amt_item = QTableWidgetItem(amt_str)
            amt_item.setForeground(QColor(amt_color))
            amt_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            amt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row_idx, 3, amt_item)
            
            # Column 4: Date
            try:
                if isinstance(r_date, str):
                    dt_obj = datetime.strptime(r_date, "%Y-%m-%d %H:%M:%S")
                else:
                    dt_obj = r_date
                formatted_date = dt_obj.strftime("%d %b %Y, %H:%M")
            except:
                formatted_date = str(r_date)
                
            date_item = QTableWidgetItem(formatted_date)
            date_item.setForeground(QColor("#94A3B8"))
            date_item.setFont(QFont("Segoe UI", 9))
            date_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_table.setItem(row_idx, 4, date_item)

        self.history_summary_lbl.setText(f"Total Records: {len(filtered_results)}  |  Total Revenue: Rs. {total_rev:,.0f}  |  Total Expenses: Rs. {total_exp:,.0f}")

    def build_reports_page(self):
        """Builds the Reports page UI."""
        layout = QVBoxLayout(self.reports_page)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)
        layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # 1. Main Card
        report_card = QFrame()
        report_card.setObjectName("ReportCard")
        report_card.setStyleSheet("""
            QFrame#ReportCard {
                background-color: white;
                border-radius: 20px;
                border: 1px solid #E2E8F0;
            }
        """)
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(20)
        shadow.setXOffset(0)
        shadow.setYOffset(4)
        shadow.setColor(QColor(0, 0, 0, 20))
        report_card.setGraphicsEffect(shadow)
        
        card_layout = QVBoxLayout(report_card)
        card_layout.setContentsMargins(30, 30, 30, 30)
        card_layout.setSpacing(16)

        # Header Title & Subtitle
        title = QLabel("📋 Generate Business Report")
        title.setStyleSheet("font-size: 20px; font-weight: 800; color: #1E293B;")
        
        subtitle = QLabel("Select a date range to generate your detailed business report")
        subtitle.setStyleSheet("font-size: 13px; color: #64748B; margin-top: -5px;")
        
        card_layout.addWidget(title)
        card_layout.addWidget(subtitle)

        # 2. Quick Select Section
        quick_label = QLabel("Quick Select")
        quick_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #1E293B; margin-top: 10px;")
        card_layout.addWidget(quick_label)
        
        quick_layout = QHBoxLayout()
        quick_layout.setSpacing(10)
        
        self.report_quick_btns = []
        quick_options = [
            ("📅 Today", 0),
            ("📅 Yesterday", 1),
            ("📅 Day Before", 2),
            ("📆 This Week", 'week'),
            ("📆 This Month", 'month')
        ]
        
        for text, key in quick_options:
            btn = QPushButton(text)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(38)
            btn.setStyleSheet(self.get_quick_btn_style(False))
            btn.clicked.connect(lambda checked, b=btn, k=key: self.handle_quick_report_date(b, k))
            quick_layout.addWidget(btn)
            self.report_quick_btns.append(btn)
            
        card_layout.addLayout(quick_layout)

        # 3. Custom Date Range Section
        custom_label = QLabel("Custom Date Range")
        custom_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #1E293B; margin-top: 16px;")
        card_layout.addWidget(custom_label)
        
        date_range_layout = QHBoxLayout()
        date_range_layout.setSpacing(20)
        
        # From Date
        from_container = QVBoxLayout()
        from_lbl = QLabel("From:")
        from_lbl.setStyleSheet("font-size: 11px; color: #64748B; margin-bottom: 2px;")
        self.report_from_date = QDateEdit()
        self.report_from_date.setCalendarPopup(True)
        self.report_from_date.setDate(QDate.currentDate())
        self.report_from_date.setFixedWidth(180)
        from_container.addWidget(from_lbl)
        from_container.addWidget(self.report_from_date)
        
        # To Date
        to_container = QVBoxLayout()
        to_lbl = QLabel("To:")
        to_lbl.setStyleSheet("font-size: 11px; color: #64748B; margin-bottom: 2px;")
        self.report_to_date = QDateEdit()
        self.report_to_date.setCalendarPopup(True)
        self.report_to_date.setDate(QDate.currentDate())
        self.report_to_date.setFixedWidth(180)
        to_container.addWidget(to_lbl)
        to_container.addWidget(self.report_to_date)
        
        date_range_layout.addLayout(from_container)
        date_range_layout.addLayout(to_container)
        date_range_layout.addStretch()
        
        # Styling for DateEdits
        date_edit_style = """
            QDateEdit {
                background-color: white;
                border: 1px solid #E2E8F0;
                border-radius: 10px;
                padding: 8px 14px;
                font-size: 13px;
                color: #1E293B;
            }
            QDateEdit::drop-down {
                border: none;
                width: 20px;
            }
        """
        self.report_from_date.setStyleSheet(date_edit_style)
        self.report_to_date.setStyleSheet(date_edit_style)
        
        # Connect changes
        self.report_from_date.dateChanged.connect(self.deactivate_quick_btns)
        self.report_to_date.dateChanged.connect(self.deactivate_quick_btns)
        
        card_layout.addLayout(date_range_layout)

        # 4. Report Format Section
        format_label = QLabel("Report Format")
        format_label.setStyleSheet("font-size: 12px; font-weight: bold; color: #1E293B; margin-top: 16px;")
        card_layout.addWidget(format_label)
        
        format_layout = QHBoxLayout()
        format_layout.setSpacing(12)
        
        self.selected_report_format = "PDF"
        self.btn_pdf_format = QPushButton("📄 PDF Report")
        self.btn_excel_format = QPushButton("📊 Excel Report")
        
        for btn in [self.btn_pdf_format, self.btn_excel_format]:
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(40)
            btn.setFixedWidth(160)
        
        self.btn_pdf_format.clicked.connect(lambda: self.set_report_format("PDF"))
        self.btn_excel_format.clicked.connect(lambda: self.set_report_format("Excel"))
        
        self.update_format_btn_styles()
        
        format_layout.addWidget(self.btn_pdf_format)
        format_layout.addWidget(self.btn_excel_format)
        format_layout.addStretch()
        
        card_layout.addLayout(format_layout)

        # 5. Generate Button
        card_layout.addSpacing(30)
        
        btn_container = QHBoxLayout()
        self.generate_report_btn = QPushButton("⚡ Generate Report")
        self.generate_report_btn.setFixedSize(300, 50)
        self.generate_report_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.generate_report_btn.setStyleSheet("""
            QPushButton {
                background-color: #1B4D89;
                color: white;
                border-radius: 12px;
                font-size: 14px;
                font-weight: 700;
            }
            QPushButton:hover {
                background-color: #3A8DFF;
            }
        """)
        self.generate_report_btn.clicked.connect(self.generate_report)
        btn_container.addStretch()
        btn_container.addWidget(self.generate_report_btn)
        btn_container.addStretch()
        card_layout.addLayout(btn_container)
        
        # 6. Status Label
        self.report_status_lbl = QLabel("")
        self.report_status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.report_status_lbl.setStyleSheet("font-size: 12px; color: #64748B; margin-top: 10px;")
        card_layout.addWidget(self.report_status_lbl)

        layout.addWidget(report_card)
        layout.addStretch()

    def get_quick_btn_style(self, active):
        if active:
            return "QPushButton { background-color: #3A8DFF; color: white; border-radius: 10px; padding: 8px 16px; font-size: 12px; font-weight: 600; border: 1px solid #3A8DFF; }"
        else:
            return "QPushButton { background-color: #F8FAFC; color: #1E293B; border-radius: 10px; padding: 8px 16px; font-size: 12px; font-weight: 600; border: 1px solid #E2E8F0; } QPushButton:hover { background-color: #DBEAFE; color: #1D4ED8; border-color: #3A8DFF; }"

    def handle_quick_report_date(self, btn, key):
        for b in self.report_quick_btns:
            b.setStyleSheet(self.get_quick_btn_style(b == btn))
            
        today = QDate.currentDate()
        self.report_from_date.blockSignals(True)
        self.report_to_date.blockSignals(True)
        
        if key == 0: # Today
            self.report_from_date.setDate(today)
            self.report_to_date.setDate(today)
        elif key == 1: # Yesterday
            yest = today.addDays(-1)
            self.report_from_date.setDate(yest)
            self.report_to_date.setDate(yest)
        elif key == 2: # Day Before
            db = today.addDays(-2)
            self.report_from_date.setDate(db)
            self.report_to_date.setDate(db)
        elif key == 'week': # This Week
            mon = today.addDays(-(today.dayOfWeek() - 1))
            self.report_from_date.setDate(mon)
            self.report_to_date.setDate(today)
        elif key == 'month': # This Month
            first = QDate(today.year(), today.month(), 1)
            self.report_from_date.setDate(first)
            self.report_to_date.setDate(today)
            
        self.report_from_date.blockSignals(False)
        self.report_to_date.blockSignals(False)

    def deactivate_quick_btns(self):
        for btn in self.report_quick_btns:
            btn.setStyleSheet(self.get_quick_btn_style(False))

    def set_report_format(self, fmt):
        self.selected_report_format = fmt
        self.update_format_btn_styles()

    def update_format_btn_styles(self):
        active = "QPushButton { background-color: #1E293B; color: white; border-radius: 10px; padding: 8px 20px; font-weight: 600; }"
        inactive = "QPushButton { background-color: #F8FAFC; color: #64748B; border-radius: 10px; padding: 8px 20px; border: 1px solid #E2E8F0; }"
        self.btn_pdf_format.setStyleSheet(active if self.selected_report_format == "PDF" else inactive)
        self.btn_excel_format.setStyleSheet(active if self.selected_report_format == "Excel" else inactive)

    def generate_report(self):
        from_date = self.report_from_date.date().toPyDate()
        to_date = self.report_to_date.date().toPyDate()
        
        if from_date > to_date:
            self.report_status_lbl.setText("❌ Error: 'From' date cannot be after 'To' date")
            self.report_status_lbl.setStyleSheet("font-size: 12px; color: #EF4444; margin-top: 10px;")
            return
            
        self.report_status_lbl.setText("⏳ Generating report...")
        self.report_status_lbl.setStyleSheet("font-size: 12px; color: #F59E0B; margin-top: 10px;")
        self.generate_report_btn.setEnabled(False)
        
        self.worker = ReportWorker(from_date, to_date, self.selected_report_format)
        self.worker.report_done.connect(self.on_report_success)
        self.worker.report_failed.connect(self.on_report_error)
        self.worker.start()

    def on_report_success(self, filepath):
        self.generate_report_btn.setEnabled(True)
        self.report_status_lbl.setText(f"✅ Report generated successfully!")
        self.report_status_lbl.setStyleSheet("font-size: 12px; color: #10B981; margin-top: 10px;")
        
        import os
        try:
            os.startfile(os.path.abspath(filepath))
        except:
            pass

    def on_report_error(self, error):
        self.generate_report_btn.setEnabled(True)
        self.report_status_lbl.setText(f"❌ Error: {error}")
        self.report_status_lbl.setStyleSheet("font-size: 12px; color: #EF4444; margin-top: 10px;")

    def on_period_changed(self, text):
        self.refresh_dashboard(text)
        if hasattr(self, 'history_table') and self.stack.currentIndex() == 1:
            self.load_history_data()

    def show_notifications(self):
        pass

    def auto_refresh(self):
        current_period = self.period_dropdown.currentText()
        self.refresh_dashboard(current_period)

    def update_notification_badge(self):
        try:
            from database.db_handler import get_notifications
            notes = get_notifications()
            if hasattr(self, 'notif_badge'):
                if len(notes) > 0:
                    self.notif_badge.show()
                else:
                    self.notif_badge.hide()
        except Exception:
            pass

    def handle_logout_click(self):
        reply = QMessageBox.question(self, 'Logout', 'Are you sure you want to logout?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.logout_clicked.emit()

    def setup_stats_cards(self):
        self.cards_layout = QHBoxLayout()
        self.cards_layout.setSpacing(20)
        
        def create_finance_card(title, value, color, icon, icon_bg):
            card = QFrame()
            card.setObjectName("FinanceCard")
            card.setFixedHeight(105)
            card.setStyleSheet(f"""
                QFrame#FinanceCard {{
                    background: #FFFFFF;
                    border-radius: 20px;
                    border: none;
                    border-top: 5px solid {color};
                    padding: 10px;
                }}
            """)
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(20)
            shadow.setXOffset(0)
            shadow.setYOffset(4)
            shadow.setColor(QColor(0, 0, 0, 20)) # 0.08 alpha
            card.setGraphicsEffect(shadow)
            
            layout = QVBoxLayout(card)
            layout.setContentsMargins(0, 0, 0, 0) # padding is inside the QFrame stylesheet
            layout.setSpacing(2)
            
            top_layout = QHBoxLayout()
            title_lbl = QLabel(title)
            title_lbl.setStyleSheet("font-size: 11px; color: #64748B; font-weight: 600; background: transparent; border: none;")
            
            icon_lbl = QLabel(icon)
            icon_lbl.setFixedSize(28, 28)
            icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            icon_lbl.setStyleSheet(f"background-color: {icon_bg}; border-radius: 14px; font-size: 12px; border: none;")
            
            top_layout.addWidget(title_lbl)
            top_layout.addStretch()
            top_layout.addWidget(icon_lbl)
            
            val_lbl = QLabel(f"Rs. {value:,.0f}")
            val_lbl.setStyleSheet(f"font-size: 22px; font-weight: 900; color: {color}; background: transparent; border: none;")
            
            layout.addLayout(top_layout)
            layout.setStretch(1, 1) # Force stretch to keep layout compact
            layout.addWidget(val_lbl)
            
            return card, val_lbl

        try:
            stats = get_consolidated_stats(CURRENT_SHOP_ID, 'Today')
            rev = stats.get('total_sales', 0)
            exp = stats.get('total_expenses', 0)
        except Exception:
            rev = 0
            exp = 0

        self.rev_card, self.revenue_value_lbl = create_finance_card(
            "💰 Total Revenue", rev, "#10B981", "💰", "#D1FAE5"
        )
        
        self.exp_card, self.expense_value_lbl = create_finance_card(
            "📊 Total Expenses", exp, "#EF4444", "📊", "#FEE2E2"
        )
        
        self.cards_layout.addWidget(self.rev_card)
        self.cards_layout.addWidget(self.exp_card)
        
        self.dash_layout.addLayout(self.cards_layout)

        # --- ROW 2: Product Sales, Repair Income, Net Profit ---
        self.row2_layout = QHBoxLayout()
        self.row2_layout.setSpacing(20)

        def create_row2_card(title, value, color, icon, icon_bg, subtitle, is_profit=False):
            card = QFrame()
            card.setObjectName("Row2Card")
            card.setFixedHeight(80)
            card.setStyleSheet(f"QFrame#Row2Card {{ background: white; border-radius: 18px; border-top: 4px solid {color}; }}")
            
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(16)
            shadow.setXOffset(0)
            shadow.setYOffset(4)
            shadow.setColor(QColor(0, 0, 0, 18)) # 0.07 alpha
            card.setGraphicsEffect(shadow)
            
            layout = QVBoxLayout(card)
            layout.setContentsMargins(10, 10, 10, 10)
            layout.setSpacing(2)
            
            top_layout = QHBoxLayout()
            title_lbl = QLabel(title)
            title_lbl.setStyleSheet("font-size: 11px; color: #64748B; font-weight: 600; border: none;")
            
            i_circle = QLabel(icon)
            i_circle.setFixedSize(26, 26)
            i_circle.setAlignment(Qt.AlignmentFlag.AlignCenter)
            i_circle.setStyleSheet(f"background-color: {icon_bg}; border-radius: 13px; font-size: 12px; border: none;")
            
            top_layout.addWidget(title_lbl)
            top_layout.addStretch()
            top_layout.addWidget(i_circle)
            
            prefix = ""
            if is_profit:
                prefix = "▲  " if value >= 0 else "▼  "
            
            val_lbl = QLabel(f"{prefix}Rs. {value:,.0f}")
            val_lbl.setStyleSheet(f"font-size: 20px; font-weight: 900; color: {color}; border: none;")
            
            sub_lbl = QLabel(subtitle)
            sub_lbl.setStyleSheet("font-size: 10px; color: #94A3B8; border: none;")
            
            layout.addLayout(top_layout)
            layout.addWidget(val_lbl)
            layout.addStretch()
            layout.addWidget(sub_lbl)
            
            return card, val_lbl, i_circle

        try:
            stats = get_consolidated_stats(CURRENT_SHOP_ID, 'Today')
            rev_total = stats.get('total_sales', 0)
            rep_rev = stats.get('repair_revenue', 0)
            prod_rev = rev_total - rep_rev
            net_prof = stats.get('net_profit', 0)
        except Exception:
            prod_rev = 0
            rep_rev = 0
            net_prof = 0

        card1, self.product_sales_lbl, _ = create_row2_card(
            "🛒 Product Sales", prod_rev, "#3A8DFF", "🛒", "#DBEAFE", "From stock sales"
        )
        
        card2, self.repair_income_lbl, _ = create_row2_card(
            "🔧 Repair Income", rep_rev, "#F59E0B", "🔧", "#FEF3C7", "From repair jobs"
        )
        
        prof_color = "#10B981" if net_prof >= 0 else "#EF4444"
        prof_bg = "#D1FAE5" if net_prof >= 0 else "#FEE2E2"
        self.net_profit_card, self.net_profit_lbl, self.net_profit_icon = create_row2_card(
            "📈 Net Profit", net_prof, prof_color, "📈", prof_bg, "Revenue minus expenses", is_profit=True
        )
        
        self.row2_layout.addWidget(card1)
        self.row2_layout.addWidget(card2)
        self.row2_layout.addWidget(self.net_profit_card)
        
        self.dash_layout.addLayout(self.row2_layout)

    def setup_center_section(self):
        center_layout = QHBoxLayout()
        center_layout.setSpacing(20)
        center_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        
        # --- LEFT SIDE: Weekly Sales Trend Chart ---
        chart_frame = QFrame()
        chart_frame.setObjectName("MainContainer")
        chart_frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        chart_frame.setMinimumHeight(320)
        chart_frame.setStyleSheet("QFrame#MainContainer { background-color: #FAFBFF; border: 1px solid #E0E0E0; border-radius: 20px; }")
        chart_layout = QVBoxLayout(chart_frame)
        chart_layout.setContentsMargins(25, 25, 25, 25)
        self.sales_chart_title = QLabel("📈 Sales & Revenue Analytics")
        self.sales_chart_title.setStyleSheet("""
            font-size: 20px; 
            font-weight: 800; 
            color: #1E293B; 
            font-family: 'Segoe UI Semibold', sans-serif;
            margin-bottom: 5px;
        """)
        chart_layout.addWidget(self.sales_chart_title)

        
        self.chart_figure = Figure(figsize=(5, 2.5), dpi=100)
        self.chart_canvas = FigureCanvas(self.chart_figure)
        self.chart_axes = self.chart_figure.add_subplot(111)
        chart_layout.addWidget(self.chart_canvas)
        
        center_layout.addWidget(chart_frame, 2)
        
        # --- RIGHT SIDE: Alerts ---
        alert_frame = QFrame()
        alert_frame.setFixedWidth(380)
        alert_frame.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        alert_frame.setMinimumHeight(320)
        alert_frame.setObjectName("MainContainer")
        alert_frame.setStyleSheet("QFrame#MainContainer { background-color: #FFFFFF; border: 1px solid #E0E0E0; border-radius: 25px; }")
        al_layout = QVBoxLayout(alert_frame)
        al_layout.setContentsMargins(25, 25, 25, 25)
        
        al_title = QLabel("📢 Recent Alerts")
        al_title.setStyleSheet("font-size: 18px; font-weight: bold; color: #1E293B; margin-bottom: 10px;")
        al_layout.addWidget(al_title)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea { background: transparent; border: none; }
            QScrollBar:vertical {
                border: none;
                background: transparent;
                width: 6px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #D1D9E6;
                border-radius: 3px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background: #3A8DFF;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
        """)
        
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        self.alerts_layout = QVBoxLayout(scroll_content)
        self.alerts_layout.setContentsMargins(0, 0, 0, 0)
        self.alerts_layout.setSpacing(10)
        
        scroll.setWidget(scroll_content)
        al_layout.addWidget(scroll, 1)
        
        center_layout.addWidget(alert_frame, 1)
        self.dash_layout.addLayout(center_layout)

    def setup_bottom_bar(self):
        bottom_row = QHBoxLayout()
        bottom_row.setContentsMargins(0, 10, 20, 20)
        
        self.expense_btn = QPushButton("➕ Add Shop Expense")
        self.expense_btn.setFixedSize(260, 50)
        self.expense_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #EF4444;
                color: white;
                border-radius: 12px;
                font-weight: bold;
                font-size: 15px;
            }
            QPushButton:hover { background-color: #DC2626; }
        """)
        self.expense_btn.clicked.connect(self.open_expense_popup)
        
        bottom_row.addStretch()
        bottom_row.addWidget(self.expense_btn)
        self.dash_layout.addLayout(bottom_row)

    def open_expense_popup(self):
        diag = ExpenseDialog(self)
        diag.data_updated.connect(lambda: self.refresh_all_dashboard_data(self.period_dropdown.currentText()))
        diag.exec()

    def handle_shop_change(self, index):
        new_id = index + 1
        set_shop_id(new_id)
        self.refresh_all_dashboard_data(self.period_dropdown.currentText())
        


    def refresh_financials(self, period):
        """Data update with key validation safety."""
        stats = get_consolidated_stats(CURRENT_SHOP_ID, period)
        
        # Update Main High-Level Stats
        mapping = {
            "Total Sales": "Total Sales",
            "Net Profit": "Net Profit",
            "Total Expenses": "Total Expenses"
        }
        
        for ui_key, stat_key in mapping.items():
            if ui_key in self.stat_widgets:
                val = stats.get(stat_key.lower().replace(' ', '_'), 0)
                self.stat_widgets[ui_key].setText(f"Rs. {val:,.0f}")
        
        # Additional Split Stats with safety checks
        if hasattr(self, 'pending_val'):
            self.pending_val.setText(str(stats.get('pending_repairs', 0)))
        if hasattr(self, 'completed_val'):
            self.completed_val.setText(str(stats.get('completed_repairs', 0)))
        
        # Sidebar Breakdown Stats
        if hasattr(self, 'op_stock_val'):
            self.op_stock_val.setText(f"Rs. {stats.get('stock_expense', 0):,.0f}")
        if hasattr(self, 'op_repair_exp_val'):
            self.op_repair_exp_val.setText(f"Rs. {stats.get('repair_expense', 0):,.0f}")
        if hasattr(self, 'op_repair_rev_val'):
            self.op_repair_rev_val.setText(f"Rs. {stats.get('repair_revenue', 0):,.0f}")

    def refresh_top_selling_items(self, period):
        # Clear top items list
        while self.items_list_layout.count():
            item = self.items_list_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
            
        top_items = get_top_items(period, 12)
        for name, qty in top_items:
            item_frame = QFrame()
            item_frame.setFixedHeight(45)
            item_frame.setStyleSheet("""
                QFrame { background-color: #F8F9FA; border-radius: 8px; border: 1px solid #E0E0E0; }
            """)
            if_layout = QHBoxLayout(item_frame)
            if_layout.setContentsMargins(15, 5, 15, 5)
            name_lbl = QLabel(name)
            name_lbl.setStyleSheet("font-weight: bold; color: #2D3436; font-size: 12px;")
            qty_lbl = QLabel(f"{qty} SOLD")
            qty_lbl.setStyleSheet("color: #10B981; font-weight: 800; font-size: 11px;")
            if_layout.addWidget(name_lbl)
            if_layout.addStretch()
            if_layout.addWidget(qty_lbl)
            self.items_list_layout.addWidget(item_frame)
        self.items_list_layout.addStretch()

    def refresh_history_table(self, period):
        self.history_table.setRowCount(0)
        data = get_history_data(period)
        
        for row_data in data:
            row_idx = self.history_table.rowCount()
            self.history_table.insertRow(row_idx)
            
            date_str = str(row_data[0]) if row_data[0] else ""
            type_str = str(row_data[1])
            desc_str = str(row_data[2])
            amt_val = float(row_data[3] or 0)
            
            self.history_table.setItem(row_idx, 0, QTableWidgetItem(date_str))
            
            type_item = QTableWidgetItem(type_str)
            if type_str == 'Sale': type_item.setForeground(QColor("#3B82F6"))
            elif type_str == 'Repair': type_item.setForeground(QColor("#F59E0B"))
            elif type_str == 'Expense': type_item.setForeground(QColor("#EF4444"))
            self.history_table.setItem(row_idx, 1, type_item)
            
            self.history_table.setItem(row_idx, 2, QTableWidgetItem(desc_str))
            
            amt_item = QTableWidgetItem(f"Rs. {amt_val:,.0f}")
            if type_str == 'Expense': amt_item.setForeground(QColor("#EF4444"))
            else: amt_item.setForeground(QColor("#10B981"))
            self.history_table.setItem(row_idx, 3, amt_item)

    def clear_dashboard(self):
        self.history_table.setRowCount(0)
        self.op_stock_val.setText("Rs. 0")
        self.op_repair_exp_val.setText("Rs. 0")
        self.op_repair_rev_val.setText("Rs. 0")
        for label in self.stat_widgets.values():
            label.setText("Rs. 0")
            
        while self.items_list_layout.count():
            item = self.items_list_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

    def load_report_data(self):
        """Prepares monthly trend data for Export purposes."""
        try:
            sales_df, expense_df, cost_df = get_analytics_data(12) # Global 12 months for report
            
            sales_df['Timestamp'] = pd.to_datetime(sales_df['Timestamp'])
            expense_df['Timestamp'] = pd.to_datetime(expense_df['Timestamp'])
            cost_df['Timestamp'] = pd.to_datetime(cost_df['Timestamp'])
            
            monthly_sales = sales_df.set_index('Timestamp')['Total_Amount'].resample('ME').sum()
            monthly_shop_exp = expense_df.set_index('Timestamp')['Amount'].resample('ME').sum()
            monthly_prod_costs = cost_df.set_index('Timestamp')['Purchase_Price'].resample('ME').sum()
            
            analytics_df = pd.DataFrame(index=monthly_sales.index)
            analytics_df['Month'] = analytics_df.index.strftime('%B %Y')
            analytics_df['Total Sales'] = monthly_sales
            analytics_df['Expenses'] = monthly_shop_exp.add(monthly_prod_costs, fill_value=0)
            analytics_df['Net Profit'] = analytics_df['Total Sales'] - analytics_df['Expenses']
            
            return analytics_df[['Month', 'Total Sales', 'Expenses', 'Net Profit']].iloc[::-1].fillna(0)
        except Exception:
            return pd.DataFrame()

    def handle_export(self):
        report_df = self.load_report_data()
        if report_df.empty:
            QMessageBox.warning(self, "No Data", "There is no performance data available to export.")
            return
            
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Performance Report", "Monthly_Security_Report.xlsx", "Excel Files (*.xlsx)"
        )
        
        if filename:
            try:
                self.export_data_to_excel(report_df, filename)
                QMessageBox.information(self, "Export Successful", f"Report saved to:\n{filename}")
                
                try:
                    from datetime import datetime
                    pdf_path = generate_daily_report(datetime.now().strftime("%Y-%m-%d"))
                    QMessageBox.information(self, "PDF Report", f"PDF Report also saved to: {pdf_path}")
                except Exception as e:
                    print(f"PDF Gen failed: {e}")
                    
            except Exception as e:
                QMessageBox.critical(self, "Export Failed", f"Error saving report: {e}")


    def export_data_to_excel(self, data_frame, filename):
        date_now_str = __import__('datetime').datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        from openpyxl.styles import Font, PatternFill, Alignment
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            data_frame.to_excel(writer, index=False, sheet_name='Performance Report', startrow=3)
            
            worksheet = writer.sheets['Performance Report']
            
            # --- 1. Headers ---
            worksheet['A1'] = f"Shop ID: {CURRENT_SHOP_ID} - Performance Report"
            worksheet['A1'].font = Font(size=14, bold=True, color="1B4D89")
            worksheet['A2'] = f"Generated On: {date_now_str}"
            worksheet['A2'].font = Font(size=10, italic=True)
            
            # --- 2. Styling Data Headers ---
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for cell in worksheet[4]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                
            # --- 3. Total Row ---
            num_rows = len(data_frame)
            if num_rows > 0:
                last_row = 4 + num_rows + 1
                worksheet.cell(row=last_row, column=1, value="TOTAL:")
                worksheet.cell(row=last_row, column=1).font = Font(bold=True)
                
                # Sum columns (excluding Month which is index 0)
                # Sales is column 2, Expenses 3, Net Profit 4
                for col_idx in [2, 3, 4]:
                    col_name = data_frame.columns[col_idx-1]
                    total_val = data_frame[col_name].sum()
                    cell = worksheet.cell(row=last_row, column=col_idx, value=total_val)
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0.00'
            
            # --- 4. Auto-width ---
            for i, col in enumerate(data_frame.columns):
                max_len = data_frame[col].astype(str).map(len).max() if not data_frame.empty else 0
                header_len = len(str(col))
                final_len = max(max_len, header_len) + 5
                worksheet.column_dimensions[chr(65 + i)].width = final_len

    def logout(self):
        self.logout_clicked.emit()
        self.close()

    def auto_logout(self):
        self.logout()

    def reset_timer(self):
        pass

    # Event overrides to detect activity
    def mouseMoveEvent(self, event):
        super().mouseMoveEvent(event)

    def keyPressEvent(self, event):
        super().keyPressEvent(event)

    def mousePressEvent(self, event):
        super().mousePressEvent(event)

    def open_settings(self):
        self.switch_view(2, self.nav_settings)

    def show_dashboard(self):
        self.stack.setCurrentIndex(0)
        self.refresh_all_dashboard_data(self.period_dropdown.currentText())

if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = AdminDashboard()
    win.show()
    sys.exit(app.exec())
