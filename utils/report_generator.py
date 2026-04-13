import os
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from database.db_handler import get_app_settings

def generate_pdf_report(data, from_date, to_date):
    """Generates a professional, premium business report in PDF format."""
    if not os.path.exists("reports"):
        os.makedirs("reports")
        
    filename = f"reports/Report_{from_date}_to_{to_date}.pdf"
    
    # Page Setup
    doc = SimpleDocTemplate(
        filename, 
        pagesize=A4, 
        rightMargin=40, 
        leftMargin=40, 
        topMargin=50, 
        bottomMargin=60 # Extra bottom margin for footer
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # Colors
    PRIMARY_BLUE = colors.HexColor('#1B4D89')
    TEXT_DARK = colors.HexColor('#1E293B')
    TEXT_MUTED = colors.HexColor('#64748B')
    BORDER_LIGHT = colors.HexColor('#E2E8F0')
    ALT_ROW = colors.HexColor('#F8FAFF')
    RED_THEME = colors.HexColor('#EF4444')
    GREEN_THEME = colors.HexColor('#10B981')
    AMBER_THEME = colors.HexColor('#F59E0B')
    TEAL_THEME = colors.HexColor('#0891B2')

    shop_settings = get_app_settings()
    shop_name = shop_settings.get("shop_name", "Kamran Electronics")
    generated_at = datetime.now().strftime("%d %b %Y, %H:%M")

    # Helper: Create Section Title
    def add_section_title(story, title, color=PRIMARY_BLUE):
        story.append(Spacer(1, 10))
        title_style = ParagraphStyle(
            'SectionTitle',
            fontSize=11,
            fontWeight='BOLD',
            textColor=color,
            textTransform='uppercase',
            spaceAfter=2
        )
        story.append(Paragraph(f"<b>{title}</b>", title_style))
        # Draw underline
        story.append(Table([[ ' ' ]], colWidths=[515], rowHeights=[1], 
                           style=[('LINEBELOW', (0,0), (-1,-1), 1, color), ('BOTTOMPADDING', (0,0), (-1,-1), 0)]))
        story.append(Spacer(1, 12))

    # --- 1. Header (First Page) ---
    def create_header_table():
        title_p = Paragraph(f"<font color='white' size='20'><b>{shop_name.upper()}</b></font>", styles['Normal'])
        subtitle_p = Paragraph(f"<font color='white' size='12'>Business Performance Report</font>", styles['Normal'])
        date_p = Paragraph(f"<font color='white' size='10'>From: {from_date}  —  To: {to_date}</font>", styles['Normal'])
        
        # Nested table for alignment
        content_table = Table([
            [title_p],
            [subtitle_p],
            [date_p]
        ], colWidths=[515])
        content_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        timestamp_p = Paragraph(f"<i>Generated: {generated_at}</i>", ParagraphStyle('ts', fontSize=8, textColor=colors.white, alignment=2))
        
        header_table = Table([
            [content_table],
            [timestamp_p]
        ], colWidths=[515], rowHeights=[65, 15])
        
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), PRIMARY_BLUE),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (1,1), (1,1), 5),
            ('RIGHTPADDING', (1,1), (1,1), 10),
        ]))
        return header_table

    story.append(create_header_table())
    story.append(Spacer(1, 25))

    # --- 2. Summary Overview ---
    add_section_title(story, "Summary Overview")
    
    summary = data['summary']
    
    def create_summary_box(label, value, color):
        # 10% opacity hex manual approx (e.g. 1A for 10% of FF)
        bg = colors.HexColor(color.hexval() + '1A') 
        
        box_inner = Table([
            [Paragraph(f"<font color='#64748B' size='8'>{label}</font>", styles['Normal'])],
            [Paragraph(f"<font color='{color.hexval()}' size='14'><b>{value}</b></font>", styles['Normal'])]
        ], colWidths=[155])
        
        box_inner.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), bg),
            ('LINEBEFORE', (0,0), (0,-1), 4, color),
            ('LEFTPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ]))
        return box_inner

    summary_grid = [
        [
            create_summary_box("TOTAL REVENUE", f"Rs. {summary.get('total_revenue', 0):,.0f}", GREEN_THEME),
            create_summary_box("TOTAL EXPENSES", f"Rs. {summary.get('total_expenses', 0):,.0f}", RED_THEME),
            create_summary_box("NET PROFIT", f"Rs. {summary.get('net_profit', 0):,.0f}", PRIMARY_BLUE)
        ],
        [
            create_summary_box("TOTAL TRANSACTIONS", str(summary.get('total_transactions', 0)), TEXT_MUTED),
            create_summary_box("TOTAL REPAIRS", str(summary.get('total_repairs', 0)), AMBER_THEME),
            create_summary_box("STOCK ADDED", str(summary.get('total_stock_added', 0)), TEAL_THEME)
        ]
    ]
    
    summary_table = Table(summary_grid, colWidths=[171, 171, 171], rowHeights=[45, 45])
    summary_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))

    # --- 3. Data Tables ---
    def create_data_table(headers, rows, col_widths=None, theme_color=PRIMARY_BLUE):
        if not rows:
            rows = [["No records for this period" + " " * (len(headers)-1)*5]] # Placeholder padding hack
            t_data = [headers] + rows
            t = Table(t_data, colWidths=col_widths or [None]*len(headers), repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), theme_color),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_MUTED),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Oblique'),
                ('GRID', (0,0), (-1,-1), 0.5, BORDER_LIGHT),
                ('SPAN', (0,1), (-1,1)),
            ]))
            return t

        t_data = [[Paragraph(f"<b>{h.upper()}</b>", ParagraphStyle('h', fontSize=9, textColor=colors.white)) for h in headers]] + rows
        t = Table(t_data, colWidths=col_widths or [None]*len(headers), repeatRows=1)
        
        style_list = [
            ('BACKGROUND', (0, 0), (-1, 0), theme_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'), # Default left
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, BORDER_LIGHT),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ALT_ROW]),
        ]
        
        # Apply special alignments
        for i, h in enumerate(headers):
            h_clean = h.upper()
            if any(term in h_clean for term in ["AMOUNT", "COST", "PRICE", "TOTAL"]):
                style_list.append(('ALIGN', (i, 1), (i, -1), 'RIGHT'))
            elif any(term in h_clean for term in ["ID", "QTY", "COUNT"]):
                style_list.append(('ALIGN', (i, 1), (i, -1), 'CENTER'))

        if theme_color == RED_THEME: # Low Stock Special
            style_list.append(('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FEF2F2')))
            style_list.append(('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#991B1B')))

        t.setStyle(TableStyle(style_list))
        return t

    # Sales
    add_section_title(story, "Sales Transactions")
    story.append(create_data_table(
        ["Date", "Customer", "Description", "Amount"], 
        [[str(r[0]), str(r[1]), str(r[2]), f"Rs. {r[3]:,.2f}"] for r in data['sales']],
        col_widths=[110, 120, 185, 100]
    ))
    
    # Repairs
    add_section_title(story, "Repair Jobs")
    story.append(create_data_table(
        ["Customer", "Item", "Issue", "Est. Cost", "Final Cost", "Status"],
        [[str(r[0]), str(r[1]), str(r[2]), str(r[3]), str(r[4]), str(r[5])] for r in data['repairs']],
        col_widths=[85, 90, 120, 70, 70, 80]
    ))

    # Expenses
    add_section_title(story, "Shop Expenses")
    story.append(create_data_table(
        ["Date", "Category", "Description", "Amount"],
        [[str(r[0]), str(r[1]), str(r[2]), f"Rs. {r[3]:,.2f}"] for r in data['expenses']],
        col_widths=[110, 110, 195, 100]
    ))

    # Stock Added
    add_section_title(story, "Stock Additions")
    story.append(create_data_table(
        ["Product Name", "Category", "Qty", "Purchase", "Selling"],
        [[str(r[0]), str(r[1]), str(r[2]), str(r[3]), str(r[4])] for r in data['stock']],
        col_widths=[180, 115, 60, 80, 80]
    ))

    # Low Stock
    if data['low_stock']:
        add_section_title(story, "Low Stock Alerts", color=RED_THEME)
        story.append(create_data_table(
            ["Product Name", "Current Stock", "Minimum Limit"],
            [[str(r[0]), str(r[1]), str(r[2])] for r in data['low_stock']],
            col_widths=[250, 140, 125],
            theme_color=RED_THEME
        ))

    # Footer Setup
    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setStrokeColor(BORDER_LIGHT)
        canvas.line(40, 45, 555, 45) # Separator
        
        # Left: Shop Name
        canvas.setFillColor(TEXT_MUTED)
        canvas.drawString(40, 32, shop_name)
        
        # Center: Confidential
        canvas.drawCentredString(297, 32, "Confidential - For Internal Use Only")
        
        # Right: Page Number
        page_num = f"Page {doc.page}"
        canvas.drawRightString(555, 32, page_num)
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return filename


def generate_excel_report(data, from_date, to_date):
    """Generates a professional Excel report using openpyxl."""
    if not os.path.exists("reports"):
        os.makedirs("reports")
        
    filename = f"reports/Report_{from_date}_to_{to_date}.xlsx"
    wb = Workbook()
    
    # Style definitions
    header_fill = PatternFill(start_color="1B4D89", end_color="1B4D89", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    center_align = Alignment(horizontal="center")

    def format_sheet(ws, title, headers):
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = max_length + 5

    # 1. Summary Sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.merge_cells('A1:C1')
    ws_sum['A1'] = f"{get_app_settings().get('shop_name', 'Kamran Electronics')} Report"
    ws_sum['A1'].font = Font(size=14, bold=True)
    ws_sum['A2'] = f"Range: {from_date} to {to_date}"
    
    summary = data['summary']
    s_rows = [
        ["Stat", "Value"],
        ["Total Revenue", f"Rs. {summary.get('total_revenue', 0):,.2f}"],
        ["Total Expenses", f"Rs. {summary.get('total_expenses', 0):,.2f}"],
        ["Net Profit", f"Rs. {summary.get('net_profit', 0):,.2f}"],
        ["Total Transactions", summary.get('total_transactions', 0)],
        ["Total Repairs", summary.get('total_repairs', 0)],
        ["Stock Added Items", summary.get('total_stock_added', 0)]
    ]
    for r in s_rows: ws_sum.append(r)

    # 2. Sales
    ws_sales = wb.create_sheet("Sales")
    format_sheet(ws_sales, "Sales Transactions", ["Date", "Customer", "Item", "Amount"])
    for r in data['sales']: ws_sales.append(r)

    # 3. Repairs
    ws_rep = wb.create_sheet("Repairs")
    format_sheet(ws_rep, "Repair Jobs", ["Customer", "Item", "Issue", "Est. Cost", "Final Cost", "Status"])
    for r in data['repairs']: ws_rep.append(r)

    # 4. Expenses
    ws_exp = wb.create_sheet("Expenses")
    format_sheet(ws_exp, "Expenses", ["Date", "Category", "Description", "Amount"])
    for r in data['expenses']: ws_exp.append(r)

    # 5. Stock
    ws_stock = wb.create_sheet("Stock")
    format_sheet(ws_stock, "Stock Additions", ["Name", "Category", "Qty", "Purchase", "Selling"])
    for r in data['stock']: ws_stock.append(r)

    # 6. Low Stock
    ws_ls = wb.create_sheet("Low Stock")
    format_sheet(ws_ls, "Low Stock Alerts", ["Product Name", "Qty", "Min Limit"])
    for r in data['low_stock']: ws_ls.append(r)

    wb.save(filename)
    return filename
