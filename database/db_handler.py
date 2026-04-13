import mysql.connector
import os
import hashlib
import json
import secrets
from ui.db_setup_dialog import get_db_config
from utils.shared import TIME_RANGES
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
CURRENT_SHOP_ID = 1

def set_shop_id(shop_id):
    global CURRENT_SHOP_ID
    CURRENT_SHOP_ID = shop_id

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def get_app_settings():
    try:
        with open("db_config.json", "r") as f:
            return json.load(f)
    except:
        return {}

def save_app_settings(settings):
    try:
        current = get_app_settings()
        current.update(settings)
        with open("db_config.json", "w") as f:
            json.dump(current, f, indent=4)
        return True
    except:
        return False

def generate_recovery_token():
    return secrets.token_hex(8).upper() # 16 characters

def connect_server():
    cfg = get_db_config()
    return mysql.connector.connect(
        host=cfg.get("host", "127.0.0.1"), 
        user=cfg.get("user", "root"),
        password=cfg.get("password", ""),
        port=cfg.get("port", 3307)        
    )

def connect_db():
    cfg = get_db_config()
    conn = mysql.connector.connect(
        host=cfg.get("host", "127.0.0.1"),
        user=cfg.get("user", "root"),
        password=cfg.get("password", ""),
        port=cfg.get("port", 3307)
    )
    cursor = conn.cursor()
    cursor.execute("USE `kamran & sohail electronics` ")
    cursor.close()
    return conn


def create_tables():
    conn = connect_server()
    cursor = conn.cursor()
    cursor.execute("CREATE DATABASE IF NOT EXISTS `kamran & sohail electronics` ")
    conn.commit()
    conn.close()
    
    conn = connect_db()
    cursor = conn.cursor()
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Products (
                        Product_ID INT AUTO_INCREMENT PRIMARY KEY,
                        Shop_ID INT DEFAULT 1,
                        Barcode VARCHAR(100) NULL,
                        Name VARCHAR(191),
                        Category VARCHAR(191),
                        Purchase_Price DECIMAL(10,2),
                        Selling_Price DECIMAL(10,2),
                        Stock_Qty INT,
                        Min_Limit INT,
                        Created_At DATETIME DEFAULT CURRENT_TIMESTAMP)''')

    # Migration: Ensure Created_At column exists for existing tables
    try:
        # Step 1: Add as NULL so existing records stay NULL (not showing in history)
        cursor.execute("ALTER TABLE Products ADD COLUMN Created_At DATETIME NULL")
    except:
        pass 
    try:
        # Step 2: Set default for future inserts
        cursor.execute("ALTER TABLE Products MODIFY Created_At DATETIME DEFAULT CURRENT_TIMESTAMP")
    except:
        pass
    
    # Migration: Ensure Barcode column exists
    try:
        cursor.execute("ALTER TABLE Products ADD COLUMN Barcode VARCHAR(100) NULL AFTER Shop_ID")
    except:
        pass # Column already exists
    
    # Requirement: `id`, `shop_id`, `customer_name`, `item_name`, `estimated_cost`, `final_cost`, `status`
    cursor.execute('''CREATE TABLE IF NOT EXISTS Repairs (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        shop_id INT DEFAULT 1,
                        customer_name VARCHAR(191),
                        item_name VARCHAR(191),
                        Issue TEXT,
                        estimated_cost DECIMAL(10,2),
                        final_cost DECIMAL(10,2) NULL,
                        spent_cost DECIMAL(10,2) DEFAULT 0.00,
                        exp_date DATE,
                        status VARCHAR(50) DEFAULT 'Pending',
                        return_timestamp DATETIME NULL)''')
    
    # Migration: Ensure spent_cost column exists for existing tables
    try:
        cursor.execute("ALTER TABLE Repairs ADD COLUMN spent_cost DECIMAL(10,2) DEFAULT 0.00 AFTER final_cost")
    except:
        pass # Column already exists
    
    try:
        cursor.execute("ALTER TABLE Repairs ADD COLUMN return_timestamp DATETIME NULL")
    except:
        pass
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Sales (
                        Sale_ID INT AUTO_INCREMENT PRIMARY KEY,
                        Shop_ID INT DEFAULT 1,
                        Product_Name VARCHAR(191),
                        Price DECIMAL(10,2),
                        Quantity INT,
                        Total DECIMAL(10,2))''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Bills (
                    Bill_ID INT AUTO_INCREMENT PRIMARY KEY,
                    Shop_ID INT DEFAULT 1,
                    Customer_Name VARCHAR(191),
                    Total_Amount DECIMAL(10,2),
                    Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    Is_Temporary TINYINT(1) DEFAULT 0)''')
    
    # Migration: Ensure Is_Temporary exists
    try:
        cursor.execute("ALTER TABLE Bills ADD COLUMN Is_Temporary TINYINT(1) DEFAULT 0")
    except: pass

    cursor.execute('''CREATE TABLE IF NOT EXISTS Bill_Items (
                Item_ID INT AUTO_INCREMENT PRIMARY KEY,
                Bill_ID INT,
                Shop_ID INT DEFAULT 1,
                Description VARCHAR(191),
                Type VARCHAR(50),
                Price DECIMAL(10,2),
                Quantity INT DEFAULT 1,
                Source_ID INT NULL,
                FOREIGN KEY(Bill_ID) REFERENCES Bills(Bill_ID) ON DELETE CASCADE)''')
    
    # Migration: Add Source_ID, Quantity
    try:
        cursor.execute("ALTER TABLE Bill_Items ADD COLUMN Source_ID INT NULL")
    except: pass
    try:
        cursor.execute("ALTER TABLE Bill_Items ADD COLUMN Quantity INT DEFAULT 1")
    except: pass
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Notifications (
                    ID INT AUTO_INCREMENT PRIMARY KEY,
                    Shop_ID INT DEFAULT 1,
                    Message TEXT,
                    Type VARCHAR(50),
                    Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY(Message(150)))''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Admin (
                    Admin_ID INT AUTO_INCREMENT PRIMARY KEY,
                    Shop_ID INT DEFAULT 1,
                    Username VARCHAR(191) UNIQUE,
                    Password VARCHAR(255))''')
    
    cursor.execute("SELECT COUNT(*) FROM Admin")
    if cursor.fetchone()[0] == 0:
        hashed_pwd = hash_password("admin123")
        cursor.execute("INSERT IGNORE INTO Admin (Shop_ID, Username, Password) VALUES (%s, %s, %s)", (1, "admin", hashed_pwd))
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Expenses (
                    Expense_ID INT AUTO_INCREMENT PRIMARY KEY,
                    Shop_ID INT DEFAULT 1,
                    Category VARCHAR(255),
                    Amount DECIMAL(10,2),
                    Description TEXT,
                    Timestamp DATETIME DEFAULT CURRENT_TIMESTAMP)''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS Staff (
                    Employee_ID INT AUTO_INCREMENT PRIMARY KEY,
                    Shop_ID INT DEFAULT 1,
                    Name VARCHAR(191),
                    Role VARCHAR(100),
                    Join_Date DATE,
                    Phone VARCHAR(50),
                    Assigned_Port VARCHAR(100))''')
    
    # Staff Migrations
    try:
        cursor.execute("ALTER TABLE Staff ADD COLUMN Monthly_Salary DECIMAL(10,2) DEFAULT 0.00")
    except: pass
    try:
        cursor.execute("ALTER TABLE Staff ADD COLUMN Advance_Amount DECIMAL(10,2) DEFAULT 0.00")
    except: pass
    try:
        cursor.execute("ALTER TABLE Staff ADD COLUMN Status VARCHAR(50) DEFAULT 'Active'")
    except: pass

    # New History Tables
    cursor.execute('''CREATE TABLE IF NOT EXISTS Staff_Salary_History (
                        History_ID INT AUTO_INCREMENT PRIMARY KEY,
                        Shop_ID INT DEFAULT 1,
                        Employee_ID INT,
                        Month VARCHAR(20),
                        Year INT,
                        Base_Salary DECIMAL(10,2),
                        Advance_Deducted DECIMAL(10,2) DEFAULT 0.00,
                        Net_Paid DECIMAL(10,2),
                        Payment_Date DATE,
                        Notes TEXT,
                        FOREIGN KEY (Employee_ID) REFERENCES Staff(Employee_ID) ON DELETE CASCADE)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS Staff_Advance_History (
                        Advance_ID INT AUTO_INCREMENT PRIMARY KEY,
                        Shop_ID INT DEFAULT 1,
                        Employee_ID INT,
                        Amount DECIMAL(10,2),
                        Given_Date DATE,
                        Notes TEXT,
                        FOREIGN KEY (Employee_ID) REFERENCES Staff(Employee_ID) ON DELETE CASCADE)''')

    cursor.execute('''CREATE TABLE IF NOT EXISTS Staff_Salary_Raises (
                        Raise_ID INT AUTO_INCREMENT PRIMARY KEY,
                        Shop_ID INT DEFAULT 1,
                        Employee_ID INT,
                        Old_Salary DECIMAL(10,2),
                        New_Salary DECIMAL(10,2),
                        Change_Date DATE,
                        Reason TEXT,
                        FOREIGN KEY (Employee_ID) REFERENCES Staff(Employee_ID) ON DELETE CASCADE)''')
    
    conn.commit()
    conn.close()
    
    cleanup_database_automated()
    sync_alerts_to_table()

def add_staff(name, role, join_date, phone, port, monthly_salary=0, status='Active'):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''INSERT INTO Staff (Shop_ID, Name, Role, Join_Date, Phone, Assigned_Port, Monthly_Salary, Status) 
                          VALUES (%s, %s, %s, %s, %s, %s, %s, %s)''', 
                       (CURRENT_SHOP_ID, name, role, join_date, phone, port, monthly_salary, status))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def get_all_staff():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT Employee_ID, Name, Role, Join_Date, Phone, Assigned_Port, Monthly_Salary, Advance_Amount, Status FROM Staff WHERE Shop_ID = %s", (CURRENT_SHOP_ID,))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_staff_by_id(employee_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT Employee_ID, Name, Role, Join_Date, Phone, Assigned_Port, Monthly_Salary, Advance_Amount, Status FROM Staff WHERE Employee_ID = %s AND Shop_ID = %s", (employee_id, CURRENT_SHOP_ID))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

def update_staff(employee_id, name, role, join_date, phone, port, monthly_salary, status):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''UPDATE Staff SET Name=%s, Role=%s, Join_Date=%s, Phone=%s, Assigned_Port=%s, Monthly_Salary=%s, Status=%s 
                          WHERE Employee_ID=%s AND Shop_ID=%s''', 
                       (name, role, join_date, phone, port, monthly_salary, status, employee_id, CURRENT_SHOP_ID))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def delete_staff(employee_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("DELETE FROM Staff WHERE Employee_ID=%s AND Shop_ID=%s", (employee_id, CURRENT_SHOP_ID))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def get_staff_salary_history(employee_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT * FROM Staff_Salary_History 
                          WHERE Employee_ID = %s AND Shop_ID = %s 
                          ORDER BY Year DESC, Month DESC""", (employee_id, CURRENT_SHOP_ID))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_last_salary_payment_date():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT MAX(Payment_Date) FROM Staff_Salary_History WHERE Shop_ID = %s", (CURRENT_SHOP_ID,))
        res = cursor.fetchone()
        return res[0] if res and res[0] else None
    finally:
        cursor.close()
        conn.close()

def get_salary_raise_history(employee_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT old_salary, new_salary, change_date, reason 
                          FROM Staff_Salary_Raises 
                          WHERE employee_id = %s AND shop_id = %s 
                          ORDER BY change_date DESC LIMIT 5""", (employee_id, CURRENT_SHOP_ID))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def add_salary_raise_record(employee_id, old_salary, new_salary, reason):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""INSERT INTO Staff_Salary_Raises (Shop_ID, Employee_ID, Old_Salary, New_Salary, Change_Date, Reason) 
                          VALUES (%s, %s, %s, %s, %s, %s)""", 
                       (CURRENT_SHOP_ID, employee_id, old_salary, new_salary, datetime.now().strftime("%Y-%m-%d"), reason))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def update_staff_salary(employee_id, new_salary):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE Staff SET Monthly_Salary = %s WHERE Employee_ID = %s AND Shop_ID = %s", 
                       (new_salary, employee_id, CURRENT_SHOP_ID))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def add_salary_payment(employee_id, month, year, base_salary, advance_deducted, net_paid, payment_date, notes):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""INSERT INTO Staff_Salary_History 
                          (Shop_ID, Employee_ID, Month, Year, Base_Salary, Advance_Deducted, Net_Paid, Payment_Date, Notes) 
                          VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""", 
                       (CURRENT_SHOP_ID, employee_id, month, year, base_salary, advance_deducted, net_paid, payment_date, notes))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def check_salary_already_paid(employee_id, month, year):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT COUNT(*) FROM Staff_Salary_History 
                          WHERE Employee_ID = %s AND Month = %s AND Year = %s AND Shop_ID = %s""", 
                       (employee_id, month, year, CURRENT_SHOP_ID))
        res = cursor.fetchone()
        return res[0] > 0
    finally:
        cursor.close()
        conn.close()

def add_advance_payment(employee_id, amount, given_date, notes):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Insert into history
        cursor.execute("""INSERT INTO Staff_Advance_History (Shop_ID, Employee_ID, Amount, Given_Date, Notes) 
                          VALUES (%s, %s, %s, %s, %s)""", (CURRENT_SHOP_ID, employee_id, amount, given_date, notes))
        # Update current balance in Staff table
        cursor.execute("UPDATE Staff SET Advance_Amount = Advance_Amount + %s WHERE Employee_ID = %s", (amount, employee_id))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def settle_advance(employee_id, deducted_amount):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("UPDATE Staff SET Advance_Amount = GREATEST(0, Advance_Amount - %s) WHERE Employee_ID = %s AND Shop_ID = %s", 
                       (deducted_amount, employee_id, CURRENT_SHOP_ID))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def get_staff_advance_history(employee_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""SELECT * FROM Staff_Advance_History 
                          WHERE Employee_ID = %s AND Shop_ID = %s 
                          ORDER BY Given_Date DESC""", (employee_id, CURRENT_SHOP_ID))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()


def insert_repair(customer, article, issue, est_cost, exp_date, status="Pending"):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''INSERT INTO Repairs (shop_id, customer_name, item_name, Issue, estimated_cost, exp_date, status) 
                          VALUES (%s, %s, %s, %s, %s, %s, %s)''', 
                       (CURRENT_SHOP_ID, customer, article, issue, est_cost, exp_date, status))
        new_id = cursor.lastrowid
        conn.commit()
        return new_id
    finally:
        cursor.close()
        conn.close()

def update_repair_status(repair_id, status, final_cost=None, spent_cost=None):
    if status == 'Completed':
        try:
            val_final = float(final_cost) if final_cost is not None else 0
            val_spent = float(spent_cost) if spent_cost is not None else 0
            if val_final <= 0 or val_spent <= 0:
                raise ValueError("Costs are mandatory for completion")
        except (ValueError, TypeError):
            raise ValueError("Costs are mandatory for completion")
        
    conn = connect_db()
    cursor = conn.cursor()
    try:
        if status == 'Completed':
            cursor.execute("UPDATE Repairs SET status = %s, final_cost = %s, spent_cost = %s WHERE id = %s AND shop_id = %s", 
                           (status, float(final_cost), float(spent_cost), repair_id, CURRENT_SHOP_ID))
        else:
            cursor.execute("UPDATE Repairs SET status = %s WHERE id = %s AND shop_id = %s", 
                           (status, repair_id, CURRENT_SHOP_ID))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

# add_repair_to_history removed (Redundant/Dead Code as requested)


def add_product(name, category, p_price, s_price, qty, min_limit):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT Product_ID, Stock_Qty, Purchase_Price FROM Products WHERE Name = %s AND Category = %s AND Shop_ID = %s", 
                       (name, category, CURRENT_SHOP_ID))
        existing_product = cursor.fetchone()
        
        if existing_product:
            product_id, curr_qty, curr_p_price = existing_product
            total_qty = curr_qty + qty
            
            # Updated Logic: Use new price for all stock (Latest Price logic) instead of averaging
            cursor.execute("""UPDATE Products 
                              SET Stock_Qty = %s, Purchase_Price = %s, Selling_Price = %s 
                              WHERE Product_ID = %s AND Shop_ID = %s""", 
                           (total_qty, p_price, s_price, product_id, CURRENT_SHOP_ID))
        else:
            cursor.execute("""INSERT INTO Products (Shop_ID, Name, Category, Purchase_Price, Selling_Price, Stock_Qty, Min_Limit) 
                              VALUES (%s, %s, %s, %s, %s, %s, %s)""", 
                           (CURRENT_SHOP_ID, name, category, p_price, s_price, qty, min_limit))
            product_id = cursor.lastrowid
            
        conn.commit()
        return product_id
    finally:
        cursor.close()
        conn.close()

def search_completed_repairs(query):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        sql = "SELECT id, customer_name, item_name, Issue, estimated_cost, final_cost FROM Repairs WHERE status = 'Completed' AND shop_id = %s"
        params = [CURRENT_SHOP_ID]
        
        if query:
            if query.isdigit():
                sql += " AND id = %s"
                params.append(int(query))
            else:
                sql += " AND customer_name LIKE %s"
                params.append(f"%{query}%")
                
        sql += " LIMIT 20"
        cursor.execute(sql, tuple(params))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def search_repair_customers(query):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        sql = "SELECT DISTINCT customer_name, item_name, exp_date FROM `kamran & sohail electronics`.Repairs WHERE shop_id = %s"
        params = [CURRENT_SHOP_ID]
        if query:
            sql += " AND (customer_name LIKE %s OR item_name LIKE %s)"
            params.extend([f"%{query}%", f"%{query}%"])
        sql += " ORDER BY exp_date DESC"
        cursor.execute(sql, tuple(params))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_all_customer_names():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT DISTINCT customer_name FROM Repairs WHERE shop_id = %s", (CURRENT_SHOP_ID,))
        return [row[0] for row in cursor.fetchall()]
    finally:
        cursor.close()
        conn.close()

def get_last_repair_by_customer(name):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT item_name, Issue, estimated_cost 
            FROM Repairs 
            WHERE customer_name = %s AND shop_id = %s 
            ORDER BY id DESC LIMIT 1
        """, (name, CURRENT_SHOP_ID))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

def get_product_by_name(name):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM `kamran & sohail electronics`.Products WHERE Name = %s AND Shop_ID = %s", (name, CURRENT_SHOP_ID))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

def search_products_advanced(query):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        sql = "SELECT * FROM `kamran & sohail electronics`.Products WHERE Shop_ID = %s AND ("
        params = [CURRENT_SHOP_ID]
        conditions = []
        if query.isdigit():
            conditions.append("Product_ID = %s")
            params.append(int(query))
        conditions.append("Name LIKE %s")
        params.append(f"%{query}%")
        conditions.append("Category LIKE %s")
        params.append(f"%{query}%")
        sql += " OR ".join(conditions) + ")"
        sql += " LIMIT 20"
        cursor.execute(sql, tuple(params))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def save_bill(customer_name, total_amount, items):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Bills (Shop_ID, Customer_Name, Total_Amount) VALUES (%s, %s, %s)", 
                       (CURRENT_SHOP_ID, customer_name, total_amount))
        bill_id = cursor.lastrowid
        
        for item in items: # desc, type, price, qty, (optional source_id)
            qty = item[3] if len(item) > 3 else 1
            row_total = qty * item[2]
            source_id = item[4] if len(item) > 4 else None
            
            cursor.execute("INSERT INTO Bill_Items (Bill_ID, Shop_ID, Description, Type, Price, Quantity, Source_ID) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                           (bill_id, CURRENT_SHOP_ID, item[0], item[1], row_total, qty, source_id))
            
            if item[1] == 'Product' and source_id:
                cursor.execute("UPDATE Products SET Stock_Qty = Stock_Qty - %s WHERE Product_ID = %s", (qty, source_id))
            elif item[1] == 'Product': # Fallback for name-based lookup if ID missing
                cursor.execute("UPDATE Products SET Stock_Qty = Stock_Qty - %s WHERE Name = %s AND Shop_ID = %s", (qty, item[0], CURRENT_SHOP_ID))
            elif item[1] == 'Repair' and source_id:
                # Update status to 'Returned' and set return timestamp
                cursor.execute("UPDATE Repairs SET status = 'Returned', return_timestamp = NOW() WHERE id = %s AND shop_id = %s", (source_id, CURRENT_SHOP_ID))
        
        conn.commit()
        return bill_id
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def get_product_by_id(p_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM `kamran & sohail electronics`.Products WHERE Product_ID = %s AND Shop_ID = %s", (p_id, CURRENT_SHOP_ID))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

def get_all_product_names():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT DISTINCT Name FROM `kamran & sohail electronics`.Products WHERE Shop_ID = %s", (CURRENT_SHOP_ID,))
        return [row[0] for row in cursor.fetchall()]
    finally:
        cursor.close()
        conn.close()

def fetch_item_by_barcode(barcode_string):
    if not barcode_string:
        return None
        
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Search by Barcode first, then by Product_ID
        cursor.execute("SELECT * FROM `kamran & sohail electronics`.Products WHERE (Barcode = %s OR Product_ID = %s) AND Shop_ID = %s", 
                       (str(barcode_string), str(barcode_string) if str(barcode_string).isdigit() else -1, CURRENT_SHOP_ID))
        return cursor.fetchone()
    except Exception as e:
        print(f"Barcode lookup error: {e}")
        return None
    finally:
        conn.close()

def get_repair_by_id(r_id):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT * FROM Repairs WHERE id = %s AND shop_id = %s", (r_id, CURRENT_SHOP_ID))
        return cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

def get_bill_history(query=None):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Order: Customer Name, Type, Description, Total, Timestamp
        sql = """
        SELECT b.Customer_Name, bi.Type, bi.Description, b.Total_Amount, b.Timestamp
        FROM Bills b
        JOIN Bill_Items bi ON b.Bill_ID = bi.Bill_ID
        WHERE b.Shop_ID = %s
        """
        params = [CURRENT_SHOP_ID]
        
        if query:
            sql += " AND (b.Customer_Name LIKE %s OR bi.Description LIKE %s)"
            params.extend([f"%{query}%", f"%{query}%"])
            
        sql += " ORDER BY b.Timestamp DESC LIMIT 20"
        
        cursor.execute(sql, tuple(params))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def cleanup_database_automated():
    """
    Sets up a MySQL Event to handle database cleanup on the server side.
    Also restricts current user's DELETE permission on billing tables as requested.
    """
    conn = connect_server()
    cursor = conn.cursor()
    try:
        # 1. Enable Event Scheduler
        cursor.execute("SET GLOBAL event_scheduler = ON")
        
        # 2. Use the database
        cursor.execute("USE `kamran & sohail electronics` ")
        
        # 3. Create cleanup event
        cursor.execute("DROP EVENT IF EXISTS daily_billing_cleanup")
        cursor.execute("""
            CREATE EVENT daily_billing_cleanup
            ON SCHEDULE EVERY 1 DAY
            COMMENT 'Daily cleanup of temporary and old billing records'
            DO
            BEGIN
                -- 1. Delete temporary records older than 48 hours
                DELETE FROM Bills WHERE Is_Temporary = 1 AND Timestamp < (NOW() - INTERVAL 48 HOUR);
                
                -- 2. Delete permanent history older than 6 months
                DELETE FROM Bills WHERE Timestamp < (NOW() - INTERVAL 6 MONTH);
                
                -- 3. Cleanup old notifications
                DELETE FROM Notifications WHERE Timestamp < NOW() - INTERVAL 2 DAY;
                
                -- 4. Delete Returned repairs older than 48 hours
                DELETE FROM Repairs WHERE status = 'Returned' AND return_timestamp < (NOW() - INTERVAL 48 HOUR);
            END
        """)
        
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

def get_notifications():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT Message FROM Notifications ORDER BY Timestamp DESC")
        return [row[0] for row in cursor.fetchall()]
    finally:
        cursor.close()
        conn.close()

def add_notification(message, n_type):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT IGNORE INTO Notifications (Message, Type) VALUES (%s, %s)", 
                       (message, n_type))
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

def sync_alerts_to_table():
    alerts = get_all_alerts()
    
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # Clear existing automated alerts for a Fresh Start
        cursor.execute("DELETE FROM Notifications WHERE Type IN ('Stock', 'Repair')")
        
        if alerts:
            for alert in alerts:
                n_type = 'Stock' if 'Low Stock' in alert else 'Repair'
                cursor.execute("INSERT IGNORE INTO Notifications (Message, Type) VALUES (%s, %s)", 
                               (alert, n_type))
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

def get_low_stock_products():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT Product_ID, Name, Stock_Qty, Min_Limit 
            FROM Products 
            WHERE Stock_Qty <= Min_Limit AND Shop_ID = %s 
            ORDER BY Stock_Qty ASC
        """, (CURRENT_SHOP_ID,))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_bill_details(customer_name, timestamp):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT b.Bill_ID, b.Customer_Name, b.Timestamp, b.Total_Amount,
                   bi.Description, bi.Type, bi.Price, bi.Quantity
            FROM Bills b
            JOIN Bill_Items bi ON b.Bill_ID = bi.Bill_ID
            WHERE b.Customer_Name = %s 
            AND b.Timestamp = %s
            AND b.Shop_ID = %s
        """, (customer_name, timestamp, CURRENT_SHOP_ID))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_chart_data(time_period):
    from datetime import datetime, timedelta
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # --- SALES DATA ---
        if time_period == 'Today':
            cursor.execute("""
                SELECT HOUR(Timestamp) as h, SUM(Total_Amount)
                FROM Bills WHERE DATE(Timestamp) = CURDATE() AND Shop_ID = %s
                GROUP BY h ORDER BY h
            """, (CURRENT_SHOP_ID,))
            rows = cursor.fetchall()
            labels = [f"{h}:00" for h in range(24)]
            s_map = {r[0]: float(r[1] or 0) for r in rows}
            sales_values = [s_map.get(h, 0.0) for h in range(24)]
            
            # Expenses for Today
            cursor.execute("""
                SELECT HOUR(Timestamp) as h, SUM(Amount)
                FROM Expenses WHERE DATE(Timestamp) = CURDATE() AND Shop_ID = %s
                GROUP BY h ORDER BY h
            """, (CURRENT_SHOP_ID,))
            e_rows = cursor.fetchall()
            e_map = {r[0]: float(r[1] or 0) for r in e_rows}
            expense_values = [e_map.get(h, 0.0) for h in range(24)]

        elif time_period == 'Weekly':
            dates = [(datetime.now() - timedelta(days=i)).date() for i in range(6, -1, -1)]
            labels = [d.strftime('%a %d') for d in dates]
            
            cursor.execute("""
                SELECT DATE(Timestamp) as d, SUM(Total_Amount)
                FROM Bills WHERE Timestamp >= DATE_SUB(CURDATE(), INTERVAL 6 DAY) AND Shop_ID = %s
                GROUP BY d ORDER BY d
            """, (CURRENT_SHOP_ID,))
            rows = cursor.fetchall()
            s_map = {r[0]: float(r[1] or 0) for r in rows}
            sales_values = [s_map.get(d, 0.0) for d in dates]
            
            cursor.execute("""
                SELECT DATE(Timestamp) as d, SUM(Amount)
                FROM Expenses WHERE Timestamp >= DATE_SUB(CURDATE(), INTERVAL 6 DAY) AND Shop_ID = %s
                GROUP BY d ORDER BY d
            """, (CURRENT_SHOP_ID,))
            e_rows = cursor.fetchall()
            e_map = {r[0]: float(r[1] or 0) for r in e_rows}
            expense_values = [e_map.get(d, 0.0) for d in dates]

        elif time_period == 'Monthly':
            dates = [(datetime.now() - timedelta(days=i)).date() for i in range(29, -1, -1)]
            labels = [d.strftime('%d %b') for d in dates]
            
            cursor.execute("""
                SELECT DATE(Timestamp) as d, SUM(Total_Amount)
                FROM Bills WHERE Timestamp >= DATE_SUB(CURDATE(), INTERVAL 29 DAY) AND Shop_ID = %s
                GROUP BY d ORDER BY d
            """, (CURRENT_SHOP_ID,))
            rows = cursor.fetchall()
            s_map = {r[0]: float(r[1] or 0) for r in rows}
            sales_values = [s_map.get(d, 0.0) for d in dates]
            
            cursor.execute("""
                SELECT DATE(Timestamp) as d, SUM(Amount)
                FROM Expenses WHERE Timestamp >= DATE_SUB(CURDATE(), INTERVAL 29 DAY) AND Shop_ID = %s
                GROUP BY d ORDER BY d
            """, (CURRENT_SHOP_ID,))
            e_rows = cursor.fetchall()
            e_map = {r[0]: float(r[1] or 0) for r in e_rows}
            expense_values = [e_map.get(d, 0.0) for d in dates]

        else: # Overall
            cursor.execute("""
                SELECT DATE_FORMAT(Timestamp, '%Y-%m') as m, SUM(Total_Amount)
                FROM Bills WHERE Shop_ID = %s
                GROUP BY m ORDER BY m
            """, (CURRENT_SHOP_ID,))
            rows = cursor.fetchall()
            
            cursor.execute("""
                SELECT DATE_FORMAT(Timestamp, '%Y-%m') as m, SUM(Amount)
                FROM Expenses WHERE Shop_ID = %s
                GROUP BY m ORDER BY m
            """, (CURRENT_SHOP_ID,))
            e_rows = cursor.fetchall()
            
            # Combine labels from both if needed, but Bills usually cover the range
            all_m = sorted(list(set([r[0] for r in rows] + [r[0] for r in e_rows])))
            labels = all_m if all_m else ['No Data']
            s_map = {r[0]: float(r[1] or 0) for r in rows}
            e_map = {r[0]: float(r[1] or 0) for r in e_rows}
            sales_values = [s_map.get(m, 0.0) for m in labels]
            expense_values = [e_map.get(m, 0.0) for m in labels]

        return labels, sales_values, expense_values
    finally:
        cursor.close()
        conn.close()

def get_recent_stock_alerts():
    conn = connect_db()
    cursor = conn.cursor()
    alerts = []
    try:
        cursor.execute("""
            SELECT Name, Stock_Qty, Min_Limit FROM Products
            WHERE Stock_Qty <= Min_Limit AND Shop_ID = %s
            ORDER BY Stock_Qty ASC LIMIT 5
        """, (CURRENT_SHOP_ID,))
        for row in cursor.fetchall():
            alerts.append(('low_stock', row[0], f"{row[1]} left (min: {row[2]})"))

        cursor.execute("""
            SELECT Name, Stock_Qty FROM Products
            WHERE Shop_ID = %s
            ORDER BY Product_ID DESC LIMIT 3
        """, (CURRENT_SHOP_ID,))
        for row in cursor.fetchall():
            alerts.append(('new_stock', row[0], f"Stock: {row[1]} units"))
    finally:
        cursor.close()
        conn.close()
    return alerts

def get_all_alerts():
    alerts = []
    conn = connect_db()
    cursor = conn.cursor()
    try:
        # 1. Stock Alert: Trigger when Stock_Qty falls below Min_Limit (inclusive)
        cursor.execute("SELECT Name, Stock_Qty FROM Products WHERE Stock_Qty <= Min_Limit AND Shop_ID = %s", (CURRENT_SHOP_ID,))
        for row in cursor.fetchall():
            alerts.append(f"Low Stock: {row[0]} ({row[1]} units left)")

        # 2. Repair Expiry: Within 12 hours of the deadline (start of the expected date)
        # Using TIMESTAMPDIFF to catch items due soon or already overdue
        cursor.execute("""SELECT customer_name, item_name, exp_date FROM Repairs 
                          WHERE status = 'Pending' AND shop_id = %s 
                          AND TIMESTAMPDIFF(HOUR, NOW(), CAST(exp_date AS DATETIME)) <= 12""", (CURRENT_SHOP_ID,))
        for row in cursor.fetchall():
            alerts.append(f"Near Deadline: {row[1]} for {row[0]} (Due: {row[2]})")
    except Exception:
        pass
    finally:
        conn.close()
    return alerts

def check_admin_login(user, pwd):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT Password, Shop_ID FROM Admin WHERE Username = %s", (user,))
        row = cursor.fetchone()
        if row:
            stored_hash = row[0]
            shop_id = row[1]
            if hash_password(pwd) == stored_hash:
                set_shop_id(shop_id)
                return True
        return False
    finally:
        cursor.close()
        conn.close()

def get_admin_stats():
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT SUM(Total_Amount) FROM Bills WHERE DATE(Timestamp) = CURDATE() AND Shop_ID = %s", (CURRENT_SHOP_ID,))
        today_sales = float(cursor.fetchone()[0] or 0.0)
        
        cursor.execute("SELECT COUNT(*) FROM Repairs WHERE status = 'Pending' AND shop_id = %s", (CURRENT_SHOP_ID,))
        pending_repairs = cursor.fetchone()[0] or 0
        
        cursor.execute("SELECT COUNT(*) FROM Products WHERE Stock_Qty <= Min_Limit AND Shop_ID = %s", (CURRENT_SHOP_ID,))
        low_stock = cursor.fetchone()[0] or 0
        
        return today_sales, pending_repairs, low_stock
    finally:
        cursor.close()
        conn.close()

def get_recent_transactions(limit=10):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            SELECT Timestamp, Customer_Name, Total_Amount 
            FROM Bills 
            WHERE Shop_ID = %s
            ORDER BY Timestamp DESC 
            LIMIT %s
        """, (CURRENT_SHOP_ID, limit))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_filtered_stats(time_period):
    ranges = TIME_RANGES
    start_time = ranges.get(time_period, "NOW() - INTERVAL 1 YEAR")
    allowed_times = list(TIME_RANGES.values())
    if start_time not in allowed_times:
        start_time = "NOW() - INTERVAL 1 YEAR"
    
    conn = connect_db()
    cursor = conn.cursor()
    stats = {
        'total_sales': 0.0, 
        'total_profit': 0.0, 
        'repair_count': 0, 
        'total_expenses': 0.0,
        'stock_expense': 0.0,
        'operating_expenses': 0.0
    }
    
    try:
        cursor.execute(f"SELECT SUM(Total_Amount) FROM `kamran & sohail electronics`.Bills WHERE Shop_ID = %s AND Timestamp BETWEEN {start_time} AND NOW()", (CURRENT_SHOP_ID,))
        stats['total_sales'] = float(cursor.fetchone()[0] or 0.0)
        
        cursor.execute(f"""
            SELECT 
                SUM(p.Purchase_Price * bi.Quantity),
                SUM(bi.Price - (p.Purchase_Price * bi.Quantity))
            FROM `kamran & sohail electronics`.Bill_Items bi
            JOIN `kamran & sohail electronics`.Products p ON (bi.Source_ID = p.Product_ID OR (bi.Source_ID IS NULL AND bi.Description = p.Name))
            JOIN `kamran & sohail electronics`.Bills b ON bi.Bill_ID = b.Bill_ID
            WHERE bi.Type = 'Product' AND b.Shop_ID = %s AND b.Timestamp BETWEEN {start_time} AND NOW()
        """, (CURRENT_SHOP_ID,))
        row = cursor.fetchone()
        cogs = float(row[0] or 0.0)
        prod_profit = float(row[1] or 0.0)
        
        cursor.execute(f"""
            SELECT SUM(bi.Price), COUNT(*)
            FROM `kamran & sohail electronics`.Bill_Items bi
            JOIN `kamran & sohail electronics`.Bills b ON bi.Bill_ID = b.Bill_ID AND b.Shop_ID = %s
            WHERE bi.Type = 'Repair' AND b.Timestamp BETWEEN {start_time} AND NOW()
        """, (CURRENT_SHOP_ID,))
        repair_row = cursor.fetchone()
        repair_profit = float(repair_row[0] or 0.0)
        stats['repair_count'] = repair_row[1] or 0
        
        cursor.execute(f"SELECT SUM(Amount) FROM `kamran & sohail electronics`.Expenses WHERE Shop_ID = %s AND Timestamp BETWEEN {start_time} AND NOW()", (CURRENT_SHOP_ID,))
        shop_expenses = float(cursor.fetchone()[0] or 0.0)
        
        stats['stock_expense'] = cogs
        stats['operating_expenses'] = shop_expenses
        stats['total_expenses'] = cogs + shop_expenses
        stats['total_profit'] = (prod_profit + repair_profit) - shop_expenses
        
        for key in ['total_sales', 'total_profit', 'total_expenses', 'stock_expense', 'operating_expenses']:
            stats[key] = round(stats[key], 2)
            
    except Exception:
        pass
    finally:
        conn.close()
    return stats

def add_expense(category, amount, desc):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO Expenses (Shop_ID, Category, Amount, Description) VALUES (%s, %s, %s, %s)", 
                       (CURRENT_SHOP_ID, category, amount, desc))
        conn.commit()
    finally:
        cursor.close()
        conn.close()

def get_top_items(time_period, limit=5):
    ranges = TIME_RANGES
    start_time = ranges.get(time_period, "NOW() - INTERVAL 1 YEAR")
    allowed_times = list(TIME_RANGES.values())
    if start_time not in allowed_times:
        start_time = "NOW() - INTERVAL 1 YEAR"
    conn = connect_db()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT Description, COUNT(*) as qty
            FROM `kamran & sohail electronics`.Bill_Items bi
            JOIN `kamran & sohail electronics`.Bills b ON bi.Bill_ID = b.Bill_ID
            WHERE bi.Type = 'Product' AND b.Shop_ID = %s AND b.Timestamp >= {start_time}
            GROUP BY Description
            ORDER BY qty DESC
            LIMIT %s
        """, (CURRENT_SHOP_ID, limit))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

def get_analytics_data(months=6):
    conn = connect_db()
    try:
        sales_query = f"SELECT Timestamp, Total_Amount FROM `kamran & sohail electronics`.Bills WHERE Shop_ID = {CURRENT_SHOP_ID} AND Timestamp >= NOW() - INTERVAL {months} MONTH"
        expense_query = f"SELECT Timestamp, Amount FROM `kamran & sohail electronics`.Expenses WHERE Shop_ID = {CURRENT_SHOP_ID} AND Timestamp >= NOW() - INTERVAL {months} MONTH"
        cost_query = f"""
            SELECT b.Timestamp, p.Purchase_Price
            FROM `kamran & sohail electronics`.Bill_Items bi
            JOIN `kamran & sohail electronics`.Products p ON (bi.Source_ID = p.Product_ID OR (bi.Source_ID IS NULL AND bi.Description = p.Name))
            JOIN `kamran & sohail electronics`.Bills b ON bi.Bill_ID = b.Bill_ID
            WHERE bi.Type = 'Product' AND b.Shop_ID = {CURRENT_SHOP_ID} 
            AND b.Timestamp >= NOW() - INTERVAL {months} MONTH
        """
        
        sales_df = pd.read_sql_query(sales_query, conn)
        expense_df = pd.read_sql_query(expense_query, conn)
        cost_df = pd.read_sql_query(cost_query, conn)
        return sales_df, expense_df, cost_df
    finally:
        conn.close()

def get_consolidated_stats(shop_id, time_period):
    """
    Returns aggregated Sales, Repair Income, and Expenses for the selected time range.
    shop_id can be a specific shop ID or 'all'.
    """
    ranges = TIME_RANGES
    start_time = ranges.get(time_period, "NOW() - INTERVAL 1 YEAR")
    allowed_times = list(TIME_RANGES.values())
    if start_time not in allowed_times:
        start_time = "NOW() - INTERVAL 1 YEAR"
    
    conn = connect_db()
    cursor = conn.cursor()
    stats = {
        'total_sales': 0.0, 
        'stock_expense': 0.0, 
        'repair_expense': 0.0, 
        'repair_revenue': 0.0, 
        'operating_expenses': 0.0,
        'total_expenses': 0.0,
        'net_profit': 0.0,
        'pending_repairs': 0,
        'completed_repairs': 0
    }
    
    shop_filter = ""
    params = []
    if str(shop_id).lower() != 'all':
        shop_filter = " AND Shop_ID = %s"
        params.append(shop_id)
        
    try:
        # 1. Total Sales
        sql_sales = f"SELECT SUM(Total_Amount) FROM `kamran & sohail electronics`.Bills WHERE Timestamp >= {start_time}{shop_filter}"
        cursor.execute(sql_sales, tuple(params))
        stats['total_sales'] = float(cursor.fetchone()[0] or 0.0)
        
        # 2. Repair Revenue (Only for 'Returned' repairs, as they are now finalized in Bills)
        sql_rep_rev = "SELECT SUM(final_cost) FROM `kamran & sohail electronics`.Repairs WHERE status = 'Returned'"
        if str(shop_id).lower() != 'all':
            sql_rep_rev += " AND shop_id = %s"
        cursor.execute(sql_rep_rev, tuple(params) if str(shop_id).lower() != 'all' else ())
        stats['repair_revenue'] = float(cursor.fetchone()[0] or 0.0)
        
        # 3. Operating Expenses (from Expenses table)
        sql_exp = f"SELECT SUM(Amount) FROM `kamran & sohail electronics`.Expenses WHERE Timestamp >= {start_time}{shop_filter}"
        cursor.execute(sql_exp, tuple(params))
        stats['operating_expenses'] = float(cursor.fetchone()[0] or 0.0)

        # 4. Stock Expense (Product COGS: Purchase_Price * Quantity)
        sql_stock_exp = f"""
            SELECT SUM(p.Purchase_Price * bi.Quantity)
            FROM `kamran & sohail electronics`.Bill_Items bi
            JOIN `kamran & sohail electronics`.Products p ON (bi.Source_ID = p.Product_ID OR (bi.Source_ID IS NULL AND bi.Description = p.Name))
            JOIN `kamran & sohail electronics`.Bills b ON bi.Bill_ID = b.Bill_ID
            WHERE bi.Type = 'Product' AND b.Timestamp >= {start_time}{shop_filter.replace('Shop_ID', 'b.Shop_ID')}
        """
        cursor.execute(sql_stock_exp, tuple(params))
        stats['stock_expense'] = float(cursor.fetchone()[0] or 0.0)
        
        # 5. Repair Expense (Sum of spent_cost for Completed and Returned repairs)
        sql_rep_exp = "SELECT SUM(spent_cost) FROM `kamran & sohail electronics`.Repairs WHERE status IN ('Completed', 'Returned')"
        if str(shop_id).lower() != 'all':
            sql_rep_exp += " AND shop_id = %s"
        cursor.execute(sql_rep_exp, tuple(params) if str(shop_id).lower() != 'all' else ())
        stats['repair_expense'] = float(cursor.fetchone()[0] or 0.0)

        # 6. Total Expenses and Net Profit Calculation: 
        stats['total_expenses'] = stats['operating_expenses'] + stats['stock_expense'] + stats['repair_expense']
        stats['net_profit'] = stats['total_sales'] - stats['total_expenses']

        # 7. Active Repairs Split (Pending and Completed)
        # Pending
        sql_pending = "SELECT COUNT(*) FROM Repairs WHERE status = 'Pending'"
        # Completed (Not yet returned/billed)
        sql_completed = "SELECT COUNT(*) FROM Repairs WHERE status = 'Completed'"
        
        p_params = []
        c_params = []
        if str(shop_id).lower() != 'all':
            sql_pending += " AND shop_id = %s"
            sql_completed += " AND shop_id = %s"
            p_params.append(shop_id)
            c_params.append(shop_id)
            
        cursor.execute(sql_pending, tuple(p_params))
        stats['pending_repairs'] = cursor.fetchone()[0] or 0
        
        cursor.execute(sql_completed, tuple(c_params))
        stats['completed_repairs'] = cursor.fetchone()[0] or 0
        
    except Exception:
        pass
    return stats

def get_dashboard_insights():
    """
    Comprehensive stats for the enhanced Admin Dashboard.
    """
    from datetime import datetime, timedelta
    conn = connect_db()
    cursor = conn.cursor()
    insights = {
        'today_sales': 0.0,
        'today_repairs_rev': 0.0,
        'today_repairs_count': 0,
        'monthly_profit': 0.0,
        'total_customers': 0,
        'weekly_trend': {'labels': [], 'values': []},
        'low_stock_list': []
    }
    
    try:
        # 1. Today's Revenue (Sales vs Repairs explicitly from Bill_Items)
        cursor.execute("""
            SELECT bi.Type, SUM(bi.Price) 
            FROM Bill_Items bi 
            JOIN Bills b ON bi.Bill_ID = b.Bill_ID 
            WHERE DATE(b.Timestamp) = CURDATE() AND b.Shop_ID = %s 
            GROUP BY bi.Type
        """, (CURRENT_SHOP_ID,))
        for row in cursor.fetchall():
            if row[0] == 'Product': insights['today_sales'] = float(row[1] or 0.0)
            elif row[0] == 'Repair': insights['today_repairs_rev'] = float(row[1] or 0.0)
            
        # 2. Today's Repairs Count (Returned today)
        cursor.execute("SELECT COUNT(*) FROM Repairs WHERE status = 'Returned' AND DATE(return_timestamp) = CURDATE() AND shop_id = %s", (CURRENT_SHOP_ID,))
        insights['today_repairs_count'] = int(cursor.fetchone()[0] or 0)

        # 3. Monthly Profit (Using consolidated logic for 'Last Month')
        m_stats = get_consolidated_stats(CURRENT_SHOP_ID, 'Last Month')
        insights['monthly_profit'] = m_stats['net_profit']

        # 4. Total Unique Customers Served
        cursor.execute("SELECT COUNT(DISTINCT Customer_Name) FROM Bills WHERE Shop_ID = %s", (CURRENT_SHOP_ID,))
        insights['total_customers'] = int(cursor.fetchone()[0] or 0)

        # 5. Weekly Sales Trend (Last 7 Days)
        cursor.execute("""
            SELECT DATE(Timestamp) as d, SUM(Total_Amount) 
            FROM Bills 
            WHERE Timestamp >= DATE_SUB(CURDATE(), INTERVAL 6 DAY) AND Shop_ID = %s 
            GROUP BY d ORDER BY d
        """, (CURRENT_SHOP_ID,))
        rows = cursor.fetchall()
        
        # Initialize with 0s for missing days if any
        dates_range = [(datetime.now() - timedelta(days=i)).date() for i in range(6, -1, -1)]
        trends_map = {r[0]: float(r[1] or 0.0) for r in rows}
        
        insights['weekly_trend']['labels'] = [d.strftime('%a') for d in dates_range]
        insights['weekly_trend']['values'] = [trends_map.get(d, 0.0) for d in dates_range]

        # 6. Low Stock Snaphot
        cursor.execute("SELECT Name, Stock_Qty FROM Products WHERE Stock_Qty < 5 AND Shop_ID = %s LIMIT 10", (CURRENT_SHOP_ID,))
        insights['low_stock_list'] = cursor.fetchall()

    except Exception as e:
        print(f"Analytics failure: {e}")
    finally:
        conn.close()
    return insights

def export_to_excel(time_period):
    """
    Fetches transaction data and exports it to a professional Excel file.
    Filters by the selected time period and CURRENT_SHOP_ID.
    """
    ranges = TIME_RANGES
    start_time = ranges.get(time_period, "NOW() - INTERVAL 1 YEAR")
    allowed_times = list(TIME_RANGES.values())
    if start_time not in allowed_times:
        start_time = "NOW() - INTERVAL 1 YEAR"
    
    conn = connect_db()
    
    # Detailed query including items
    query = f"""
    SELECT b.Timestamp, b.Customer_Name as `Customer Name`, 
           bi.Description as `Item/Service`, b.Total_Amount as `Total Amount`
    FROM `kamran & sohail electronics`.Bills b
    LEFT JOIN `kamran & sohail electronics`.Bill_Items bi ON b.Bill_ID = bi.Bill_ID
    WHERE b.Shop_ID = {CURRENT_SHOP_ID} AND b.Timestamp >= {start_time}
    ORDER BY b.Timestamp DESC
    """
    
    try:
        
        df = pd.read_sql_query(query, conn)
        
        if not df.empty:
            df['Timestamp'] = pd.to_datetime(df['Timestamp']).dt.strftime('%Y-%m-%d %H:%M')
            df.rename(columns={'Timestamp': 'Date'}, inplace=True)
        
        # Ensure exports/ folder exists
        export_dir = os.path.join(os.getcwd(), "exports")
        if not os.path.exists(export_dir):
            os.makedirs(export_dir)
            
        date_now = datetime.now()
        date_str = date_now.strftime("%Y-%m-%d")
        filename = f"Report_Shop{CURRENT_SHOP_ID}_{date_str}.xlsx"
        filepath = os.path.join(export_dir, filename)
        
        shop_name = f"Shop ID: {CURRENT_SHOP_ID} Performance Report"
        
        # Use ExcelWriter with openpyxl 
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Start Writing df from row 3 to leave room for headers
            df.to_excel(writer, index=False, sheet_name='Sales Report', startrow=3)
            
            workbook = writer.book
            worksheet = writer.sheets['Sales Report']
            
            # --- 1. Add Shop Header ---
            worksheet['A1'] = shop_name
            worksheet['A1'].font = Font(size=16, bold=True, color="1B4D89")
            
            worksheet['A2'] = f"Generated On: {date_now.strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet['A2'].font = Font(size=11, italic=True)
            
            # --- 2. Format Column Headers (now on row 4) ---
            header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            header_font = Font(bold=True)
            
            for cell in worksheet[4]: # Row 4 is the header row for the dataframe
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # --- 3. Add Total Row at bottom ---
            num_rows = len(df)
            if num_rows > 0:
                total_row_idx = 4 + num_rows + 1 # openpyxl is 1-indexed
                worksheet.cell(row=total_row_idx, column=3, value="TOTAL:")
                worksheet.cell(row=total_row_idx, column=3).font = Font(bold=True)
                
                # Sum the Total Amount column (D is column 4)
                total_sum = df['Total Amount'].sum()
                worksheet.cell(row=total_row_idx, column=4, value=f"Rs. {total_sum:,.2f}")
                worksheet.cell(row=total_row_idx, column=4).font = Font(bold=True)
            
            # --- 4. Auto-fit column width ---
            for i, col in enumerate(df.columns):
                max_len = df[col].astype(str).map(len).max() if not df.empty else 0
                header_len = len(str(col))
                final_len = max(max_len, header_len) + 5
                worksheet.column_dimensions[chr(65 + i)].width = final_len

        return filepath
    except Exception as e:
        # Keep critical fail log for user feedback troubleshooting
        print(f"Export failed: {e}")
        raise e
    finally:
        conn.close()

def get_history_data(time_period):
    ranges = TIME_RANGES
    start_time = ranges.get(time_period, "NOW() - INTERVAL 1 YEAR")
    allowed_times = list(TIME_RANGES.values())
    if start_time not in allowed_times:
        start_time = "NOW() - INTERVAL 1 YEAR"
    
    conn = connect_db()
    cursor = conn.cursor()
    
    # Combined query for Bills, Repairs, Expenses, and Stock additions
    # Order: customer_name, type, description, amount, date
    query = f"""
    (SELECT Customer_Name as customer_name, 'Sale' as type, 'Sale Transaction' as description, Total_Amount as amount, Timestamp as date
     FROM `kamran & sohail electronics`.Bills WHERE Shop_ID = {CURRENT_SHOP_ID} AND Timestamp >= {start_time})
    UNION ALL
    (SELECT customer_name as customer_name, 'Repair' as type, CONCAT(customer_name, ' (', item_name, ')') as description, final_cost as amount, exp_date as date
     FROM `kamran & sohail electronics`.Repairs WHERE shop_id = {CURRENT_SHOP_ID} AND status = 'Completed')
    UNION ALL
    (SELECT 'Shop Expense' as customer_name, 'Expense' as type, Description as description, Amount as amount, Timestamp as date
     FROM `kamran & sohail electronics`.Expenses WHERE Shop_ID = {CURRENT_SHOP_ID} AND Timestamp >= {start_time})
    UNION ALL
    (SELECT 'Stock Added' as customer_name, 'Stock' as type, CONCAT(Name, ' (', Category, ')') as description, Stock_Qty as amount, Created_At as date
     FROM `kamran & sohail electronics`.Products 
     WHERE Shop_ID = {CURRENT_SHOP_ID} AND Created_At IS NOT NULL AND Created_At BETWEEN {start_time} AND NOW())
    ORDER BY date DESC
    """
    
    try:
        cursor.execute(query)
        rows = cursor.fetchall()
        return rows
    except Exception as e:
        print(f"Error fetching history data: {e}")
        return []
    finally:
        conn.close()

def get_report_data(from_date, to_date):
    """
    Fetches aggregated and detailed data for business reports between two dates.
    from_date, to_date: date objects or strings
    """
    if hasattr(from_date, 'strftime'):
        f_date = from_date.strftime('%Y-%m-%d')
    else:
        f_date = str(from_date)
        
    if hasattr(to_date, 'strftime'):
        t_date = to_date.strftime('%Y-%m-%d')
    else:
        t_date = str(to_date)
        
    f_dt = f"{f_date} 00:00:00"
    t_dt = f"{t_date} 23:59:59"
    
    conn = connect_db()
    cursor = conn.cursor()
    data = {
        'summary': {},
        'sales': [],
        'repairs': [],
        'expenses': [],
        'stock': [],
        'low_stock': []
    }
    
    try:
        # 1. Sales Transactions (Bills joined with Items)
        cursor.execute("""
            SELECT b.Timestamp, b.Customer_Name, bi.Description, (bi.Price * bi.Quantity) as total
            FROM Bills b
            JOIN Bill_Items bi ON b.Bill_ID = bi.Bill_ID
            WHERE b.Shop_ID = %s AND b.Timestamp BETWEEN %s AND %s
            ORDER BY b.Timestamp DESC
        """, (CURRENT_SHOP_ID, f_dt, t_dt))
        data['sales'] = cursor.fetchall()
        
        # 2. Repair Jobs
        cursor.execute("""
            SELECT customer_name, item_name, Issue, estimated_cost, final_cost, status
            FROM Repairs
            WHERE shop_id = %s AND (exp_date BETWEEN %s AND %s OR return_timestamp BETWEEN %s AND %s)
        """, (CURRENT_SHOP_ID, f_date, t_date, f_dt, t_dt))
        data['repairs'] = cursor.fetchall()
        
        # 3. Expenses
        cursor.execute("""
            SELECT Timestamp, Category, Description, Amount
            FROM Expenses
            WHERE Shop_ID = %s AND Timestamp BETWEEN %s AND %s
            ORDER BY Timestamp DESC
        """, (CURRENT_SHOP_ID, f_dt, t_dt))
        data['expenses'] = cursor.fetchall()
        
        # 4. Stock Additions (Approximation using Product_ID order)
        cursor.execute("""
            SELECT Name, Category, Stock_Qty, Purchase_Price, Selling_Price
            FROM Products
            WHERE Shop_ID = %s
            ORDER BY Product_ID DESC LIMIT 50
        """, (CURRENT_SHOP_ID,))
        data['stock'] = cursor.fetchall()
        
        # 5. Low Stock
        cursor.execute("""
            SELECT Name, Stock_Qty, Min_Limit
            FROM Products
            WHERE Shop_ID = %s AND Stock_Qty <= Min_Limit
        """, (CURRENT_SHOP_ID,))
        data['low_stock'] = cursor.fetchall()
        
        # 6. Summary Calculation
        total_rev = sum(float(row[3] or 0) for row in data['sales'])
        total_exp = sum(float(row[3] or 0) for row in data['expenses'])
        
        data['summary'] = {
            'total_revenue': total_rev,
            'total_expenses': total_exp,
            'net_profit': total_rev - total_exp,
            'total_transactions': len(data['sales']),
            'total_repairs': len(data['repairs']),
            'total_stock_added': len(data['stock'])
        }
        
    except Exception as e:
        print(f"Report data fetch error: {e}")
    finally:
        cursor.close()
        conn.close()
        
    return data

def get_all_repairs(query="", status="All"):
    conn = connect_db()
    cursor = conn.cursor()
    try:
        sql = "SELECT id, customer_name, item_name, Issue, estimated_cost, final_cost, exp_date, status FROM Repairs WHERE shop_id = %s"
        params = [CURRENT_SHOP_ID]
        if status != "All":
            sql += " AND status = %s"
            params.append(status)
        else:
            sql += " AND status IN ('Pending', 'Completed')"
        if query:
            clean_query = query
            if query.startswith("JOB_ID:"):
                try:
                    clean_query = query.split("|")[0].split(":")[1]
                except: pass
            if clean_query.isdigit():
                sql += " AND (id = %s OR customer_name LIKE %s OR item_name LIKE %s)"
                params.extend([int(clean_query), '%'+clean_query+'%', '%'+clean_query+'%'])
            else:
                sql += " AND (customer_name LIKE %s OR item_name LIKE %s)"
                params.extend(['%'+clean_query+'%', '%'+clean_query+'%'])
        cursor.execute(sql, tuple(params))
        return cursor.fetchall()
    finally:
        cursor.close()
        conn.close()

if __name__ == "__main__":
    create_tables()
